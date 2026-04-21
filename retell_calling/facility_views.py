"""
Views for managing Facility Leads
Handles uploading facility lists, managing calls, and tracking engagement
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.db import models
from django.utils import timezone
from django.core.paginator import Paginator
from django.conf import settings

import json
import logging
import tempfile
from datetime import datetime

from .models import (
    FacilityLead, FacilityCallSession, FacilityCallSummary, 
    BulkCallSession
)
from .forms import FacilityLeadExcelUploadForm, FacilityLeadFilterForm
from .services import RetellCallService
from medications.gemini_service import GeminiSummaryService

logger = logging.getLogger('retell_calling.views')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def facility_leads_list(request):
    """Display list of all facility leads with filtering and upload capabilities"""
    try:
        # Handle Excel import if POST request
        if request.method == 'POST' and 'excel_file' in request.FILES:
            return handle_facility_excel_import(request)
        
        # Get filter parameters
        search_query = request.GET.get('search', '').strip()
        category_filter = request.GET.get('category', '').strip()
        city_filter = request.GET.get('city', '').strip()
        call_status_filter = request.GET.get('call_status', '').strip()
        email_status_filter = request.GET.get('email_status', '').strip()
        
        # Start with all facility leads
        facilities = FacilityLead.objects.all()
        
        # Apply search filter
        if search_query:
            facilities = facilities.filter(
                models.Q(facility_name__icontains=search_query) |
                models.Q(city__icontains=search_query) |
                models.Q(phone_number__icontains=search_query)
            )
        
        # Apply category filter
        if category_filter:
            facilities = facilities.filter(category__icontains=category_filter)
        
        # Apply city filter
        if city_filter:
            facilities = facilities.filter(city__icontains=city_filter)
        
        # Apply call status filter
        if call_status_filter == 'called':
            facilities = facilities.filter(call_attempted=True)
        elif call_status_filter == 'not_called':
            facilities = facilities.filter(call_attempted=False)
        
        # Apply email status filter
        if email_status_filter == 'sent':
            facilities = facilities.filter(email_sent=True)
        elif email_status_filter == 'not_sent':
            facilities = facilities.filter(email_sent=False)
        
        # Order by facility name
        facilities = facilities.order_by('facility_name')
        
        # Calculate statistics
        total_facilities = FacilityLead.objects.count()
        called_facilities = FacilityLead.objects.filter(call_attempted=True).count()
        email_sent_count = FacilityLead.objects.filter(email_sent=True).count()
        call_success_rate = (called_facilities / total_facilities * 100) if total_facilities > 0 else 0
        
        # Implement pagination
        paginator = Paginator(facilities, 25)  # Show 25 facilities per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Create upload form
        upload_form = FacilityLeadExcelUploadForm()
        filter_form = FacilityLeadFilterForm(request.GET)
        
        context = {
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': page_obj.has_other_pages(),
            'upload_form': upload_form,
            'filter_form': filter_form,
            'search_query': search_query,
            'category_filter': category_filter,
            'city_filter': city_filter,
            'call_status_filter': call_status_filter,
            'email_status_filter': email_status_filter,
            'total_facilities': total_facilities,
            'called_facilities': called_facilities,
            'email_sent_count': email_sent_count,
            'call_success_rate': round(call_success_rate, 1),
            'filtered_count': facilities.count()
        }
        
        return render(request, 'admin/facility_leads_list.html', context)
        
    except Exception as e:
        logger.error(f'Error loading facility leads list: {str(e)}')
        messages.error(request, f'Error loading facility leads: {str(e)}')
        return redirect('admin_dashboard')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def facility_lead_detail(request, facility_id):
    """Display detailed information about a specific facility lead"""
    try:
        facility = get_object_or_404(FacilityLead, id=facility_id)
        
        # Get all call sessions for this facility
        call_sessions = facility.call_sessions.all().order_by('-created_at')
        
        # Get call summaries
        call_summaries = facility.call_summaries.all().order_by('-generated_at')
        
        context = {
            'facility': facility,
            'call_sessions': call_sessions,
            'call_summaries': call_summaries,
            'has_calls': call_sessions.exists(),
            'has_summaries': call_summaries.exists()
        }
        
        return render(request, 'admin/facility_lead_detail.html', context)
        
    except Exception as e:
        logger.error(f'Error loading facility details: {str(e)}')
        messages.error(request, f'Error loading facility details: {str(e)}')
        return redirect('facility_leads_list')


def handle_facility_excel_import(request):
    """Handle Excel file import for facility leads"""
    try:
        import openpyxl
        
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            messages.error(request, 'No file uploaded.')
            return redirect('facility_leads_list')
        
        # Validate file type
        if not excel_file.name.lower().endswith(('.xlsx', '.xls', '.csv')):
            messages.error(request, 'Please upload a valid Excel or CSV file.')
            return redirect('facility_leads_list')
        
        logger.info(f"Processing facility Excel file: {excel_file.name}")
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(tmp_file_path)
            worksheet = workbook.active
            
            # Define column mappings based on Excel headers
            # Expected columns: Category, Facility Name, Phone Number, Address, City, ZIP, Capacity, Hospice, Non-Ambulatory
            header_row = worksheet[1]
            headers = [cell.value for cell in header_row]
            
            logger.info(f"Excel headers: {headers}")
            
            # Create a mapping of column names to column indices
            column_map = {}
            for idx, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    column_map[header_lower] = idx + 1  # openpyxl uses 1-based indexing
            
            # Map expected column names to actual column indices
            facility_name_col = None
            phone_col = None
            address_col = None
            city_col = None
            zip_col = None
            category_col = None
            capacity_col = None
            hospice_col = None
            non_ambulatory_col = None
            
            # Try to find columns with flexible matching
            for header, col_idx in column_map.items():
                if 'facility' in header and 'name' in header:
                    facility_name_col = col_idx
                elif 'phone' in header:
                    phone_col = col_idx
                elif 'address' in header:
                    address_col = col_idx
                elif 'city' in header:
                    city_col = col_idx
                elif 'zip' in header or 'postal' in header:
                    zip_col = col_idx
                elif 'category' in header:
                    category_col = col_idx
                elif 'capacity' in header:
                    capacity_col = col_idx
                elif 'hospice' in header:
                    hospice_col = col_idx
                elif 'non' in header and 'ambulatory' in header:
                    non_ambulatory_col = col_idx
            
            # Fallback to positional mapping if column mapping didn't work
            if not all([facility_name_col, phone_col, address_col, city_col, zip_col]):
                # Assume standard order: Category, Facility Name, Phone, Address, City, ZIP, Capacity, Hospice, Non-Ambulatory
                category_col = 1
                facility_name_col = 2
                phone_col = 3
                address_col = 4
                city_col = 5
                zip_col = 6
                capacity_col = 7
                hospice_col = 8
                non_ambulatory_col = 9
            
            # Process rows
            created_count = 0
            duplicate_count = 0
            error_count = 0
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Extract cell values
                    facility_name = worksheet.cell(row_idx, facility_name_col).value
                    phone_number = worksheet.cell(row_idx, phone_col).value
                    address = worksheet.cell(row_idx, address_col).value
                    city = worksheet.cell(row_idx, city_col).value
                    zip_code = worksheet.cell(row_idx, zip_col).value
                    
                    # Optional fields
                    category = worksheet.cell(row_idx, category_col).value if category_col else None
                    capacity = worksheet.cell(row_idx, capacity_col).value if capacity_col else None
                    hospice = worksheet.cell(row_idx, hospice_col).value if hospice_col else None
                    non_ambulatory = worksheet.cell(row_idx, non_ambulatory_col).value if non_ambulatory_col else None
                    
                    # Skip empty rows
                    if not all([facility_name, phone_number, city]):
                        continue
                    
                    # Convert capacity to integer if present
                    if capacity:
                        try:
                            capacity = int(capacity)
                        except (ValueError, TypeError):
                            capacity = None
                    
                    # Normalize phone number (remove special characters)
                    phone_number = str(phone_number).replace('-', '').replace(' ', '').replace('(', '').replace(')', '')
                    
                    # Try to create or update facility lead
                    try:
                        facility_lead, created = FacilityLead.objects.get_or_create(
                            facility_name=str(facility_name).strip(),
                            phone_number=phone_number,
                            city=str(city).strip(),
                            defaults={
                                'address': str(address).strip() if address else '',
                                'zip_code': str(zip_code).strip() if zip_code else '',
                                'category': str(category).strip() if category else None,
                                'capacity': capacity,
                                'hospice': 'Y' if str(hospice).upper() in ['Y', 'YES'] else ('N' if str(hospice).upper() in ['N', 'NO'] else None) if hospice else None,
                                'non_ambulatory': 'Y' if str(non_ambulatory).upper() in ['Y', 'YES'] else ('N' if str(non_ambulatory).upper() in ['N', 'NO'] else None) if non_ambulatory else None,
                                'uploaded_by': request.user
                            }
                        )
                        
                        if created:
                            created_count += 1
                            logger.info(f"Created facility lead: {facility_name}")
                        else:
                            duplicate_count += 1
                            logger.info(f"Facility lead already exists: {facility_name}")
                    
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Error creating facility lead at row {row_idx}: {str(e)}")
                
                except Exception as e:
                    error_count += 1
                    logger.error(f"Error processing row {row_idx}: {str(e)}")
            
            # Provide feedback to user
            messages.success(
                request,
                f'Successfully imported {created_count} facility leads. '
                f'{duplicate_count} duplicates skipped. '
                f'{error_count} errors encountered.'
            )
            
            logger.info(
                f"Facility Excel import completed: Created {created_count}, "
                f"Duplicates {duplicate_count}, Errors {error_count}"
            )
        
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            messages.error(request, f'Error processing Excel file: {str(e)}')
        
        finally:
            # Clean up temporary file
            import os
            try:
                os.remove(tmp_file_path)
            except:
                pass
        
        return redirect('facility_leads_list')
    
    except Exception as e:
        logger.error(f"Error in handle_facility_excel_import: {str(e)}")
        messages.error(request, f'Error importing facility data: {str(e)}')
        return redirect('facility_leads_list')


@api_view(['POST'])
def trigger_facility_bulk_calls(request):
    """
    Trigger sequential calls to all facilities without existing calls.
    This returns immediately and starts a background process to handle calls sequentially.
    
    Expected payload:
    {
        "agent_id": "optional_agent_id"
    }
    """
    try:
        data = request.data
        
        # Use default agent ID
        agent_id = "agent_f1e1852a2a6b90b39d997f95b5"
        
        logger.info("Starting bulk facility calling process")
        
        # Get facilities without calls
        facilities_without_calls = FacilityLead.objects.annotate(
            call_count=models.Count('call_sessions')
        ).filter(call_count=0).filter(phone_number__isnull=False).exclude(phone_number='')
        
        if not facilities_without_calls.exists():
            return Response({
                'success': True,
                'message': 'No facilities available for calling',
                'facilities_to_call': 0,
                'facilities': []
            }, status=status.HTTP_200_OK)
        
        facility_count = facilities_without_calls.count()
        facility_list = []
        
        for facility in facilities_without_calls:
            facility_list.append({
                'id': str(facility.id),
                'name': facility.facility_name,
                'phone': facility.phone_number,
                'city': facility.city,
                'address': facility.address
            })
        
        # Create a bulk calling session to track progress
        bulk_session = BulkCallSession.objects.create(
            session_type='facility_calls',
            total_leads=facility_count,
            leads_data=facility_list,
            agent_id=agent_id,
            status='initiated'
        )
        
        # Start the first call
        result = initiate_next_facility_call(bulk_session)
        
        if result['success']:
            return Response({
                'success': True,
                'message': f'Bulk calling process started for {facility_count} facilities',
                'bulk_session_id': str(bulk_session.id),
                'facilities_to_call': facility_count,
                'first_call_initiated': True,
                'first_call_id': result.get('call_id'),
                'facilities': facility_list
            }, status=status.HTTP_201_CREATED)
        else:
            bulk_session.status = 'failed'
            bulk_session.error_message = result.get('error', 'Failed to start first call')
            bulk_session.save()
            
            return Response({
                'success': False,
                'error': f'Failed to start bulk calling: {result.get("error")}',
                'bulk_session_id': str(bulk_session.id)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    except Exception as e:
        logger.error(f"Unexpected error in trigger_facility_bulk_calls: {str(e)}")
        return Response({
            'error': f'Internal server error: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def initiate_next_facility_call(bulk_session, facility=None):
    """
    Initiate the next call in a bulk calling session for facilities.
    Always uses bulk_session.current_index to determine which facility to call.
    """
    try:
        # Check if we've reached the end
        if bulk_session.current_index >= len(bulk_session.leads_data):
            logger.info(f"No more facilities to call in bulk session {bulk_session.id}")
            bulk_session.status = 'completed'
            bulk_session.completed_at = timezone.now()
            bulk_session.save()
            return {'success': False, 'message': 'No more facilities to call'}
        
        # Get facility data from the session
        facility_data = bulk_session.leads_data[bulk_session.current_index]
        logger.info(f"Attempting to call facility at index {bulk_session.current_index}: {facility_data}")
        
        try:
            current_facility = FacilityLead.objects.get(id=facility_data['id'])
        except FacilityLead.DoesNotExist:
            logger.error(f"Facility {facility_data['id']} at index {bulk_session.current_index} not found")
            bulk_session.mark_call_completed(
                success=False,
                call_data={
                    'facility_id': facility_data['id'],
                    'error': 'Facility not found',
                    'timestamp': timezone.now().isoformat(),
                    'index': bulk_session.current_index
                }
            )
            return {'success': False, 'error': 'Facility not found'}
        
        # Check if facility already has calls
        if current_facility.call_sessions.exists():
            logger.info(f"Facility {current_facility.id} already has calls")
            bulk_session.mark_call_completed(
                success=False,
                call_data={
                    'facility_id': current_facility.id,
                    'error': 'Facility already has calls',
                    'timestamp': timezone.now().isoformat(),
                    'index': bulk_session.current_index
                }
            )
            return {'success': False, 'error': 'Facility already has calls'}
        
        # Initialize call service
        call_service = RetellCallService()
        
        logger.info(f"Initiating call to facility {current_facility.id}: {current_facility.facility_name}")
        
        # Build dynamic variables
        dynamic_variables = {
            'facility_id': str(current_facility.id),
            'facility_name': current_facility.facility_name,
            'facility_phone': current_facility.phone_number,
            'facility_city': current_facility.city,
            'facility_address': current_facility.address,
            'category': current_facility.category or '',
            'capacity': str(current_facility.capacity) if current_facility.capacity else '',
        }
        
        # Create the call (using a facility-specific method if available)
        result = call_service.create_facility_call(current_facility, bulk_session.agent_id, dynamic_variables)
        
        call_session = result['call_session']
        call_id = result['call_id']
        
        # Associate with bulk session
        call_session.bulk_session_id = bulk_session.id
        call_session.save()
        
        # Mark facility as having a call attempted
        current_facility.call_attempted = True
        current_facility.save()
        
        logger.info(f"Call successfully initiated for facility {current_facility.id}, call_id: {call_id}")
        
        # Update bulk session status
        bulk_session.status = 'in_progress'
        bulk_session.save()
        
        return {
            'success': True,
            'call_id': call_id,
            'call_session_id': str(call_session.id),
            'facility_id': current_facility.id,
            'facility_name': current_facility.facility_name,
            'bulk_session_id': str(bulk_session.id),
            'current_index': bulk_session.current_index
        }
    
    except Exception as e:
        logger.error(f"Error initiating facility call at index {bulk_session.current_index}: {str(e)}")
        bulk_session.mark_call_completed(
            success=False,
            call_data={
                'facility_id': 'unknown',
                'error': str(e),
                'timestamp': timezone.now().isoformat(),
                'index': bulk_session.current_index
            }
        )
        return {'success': False, 'error': str(e)}


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def facility_call_summaries_list(request):
    """Display list of all facility call summaries"""
    try:
        from django.core.paginator import Paginator
        
        # Get all call summaries
        summaries = FacilityCallSummary.objects.all().order_by('-generated_at')
        
        # Apply search if provided
        search_query = request.GET.get('search', '').strip()
        if search_query:
            summaries = summaries.filter(
                models.Q(facility__facility_name__icontains=search_query) |
                models.Q(facility__city__icontains=search_query)
            )
        
        # Calculate statistics
        total_summaries = FacilityCallSummary.objects.count()
        recent_summaries = FacilityCallSummary.objects.filter(
            generated_at__gte=timezone.now() - timezone.timedelta(days=7)
        ).count()
        
        # Pagination
        paginator = Paginator(summaries, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': page_obj.has_other_pages(),
            'search_query': search_query,
            'total_summaries': total_summaries,
            'recent_summaries': recent_summaries,
            'filtered_count': summaries.count()
        }
        
        return render(request, 'admin/facility_call_summaries_list.html', context)
    
    except Exception as e:
        logger.error(f'Error loading facility call summaries: {str(e)}')
        messages.error(request, f'Error loading call summaries: {str(e)}')
        return redirect('admin_dashboard')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def send_facility_followup_emails(request):
    """Send follow-up emails to facilities whose calls have completed"""
    try:
        # Get all completed calls that haven't had emails sent yet
        completed_calls = FacilityCallSession.objects.filter(
            email_sent=False,
            call_status='completed'
        )
        
        if not completed_calls.exists():
            messages.info(request, 'No pending emails to send.')
            return redirect('facility_leads_list')
        
        email_count = 0
        error_count = 0
        
        for call_session in completed_calls:
            try:
                # Send email to facility (implementation below)
                send_facility_email(call_session)
                call_session.email_sent = True
                call_session.email_sent_at = timezone.now()
                call_session.save()
                
                # Update facility record
                call_session.facility.email_sent = True
                call_session.facility.save()
                
                email_count += 1
                logger.info(f"Email sent to facility: {call_session.facility.facility_name}")
            
            except Exception as e:
                error_count += 1
                logger.error(f"Error sending email to facility {call_session.facility.facility_name}: {str(e)}")
        
        messages.success(
            request,
            f'Successfully sent {email_count} emails. {error_count} errors encountered.'
        )
        
        return redirect('facility_leads_list')
    
    except Exception as e:
        logger.error(f'Error in send_facility_followup_emails: {str(e)}')
        messages.error(request, f'Error sending emails: {str(e)}')
        return redirect('facility_leads_list')


def send_facility_email(call_session):
    """
    Send follow-up email to facility with call summary and attachments
    """
    try:
        import sendgrid
        from sendgrid.helpers.mail import Mail, Attachment, ContentDisposition
        import base64
        
        facility = call_session.facility
        
        # Skip if no SendGrid API key
        if not hasattr(settings, 'SENDGRID_API_KEY') or not settings.SENDGRID_API_KEY:
            logger.warning(f"SendGrid not configured. Skipping email for facility {facility.facility_name}")
            return False
        
        # Get or create summary
        summary = call_session.summary
        
        # Build email content
        email_html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2 style="color: #7928CA;">Call Summary - {facility.facility_name}</h2>
            
            <div style="background: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <p><strong>Facility:</strong> {facility.facility_name}</p>
                <p><strong>Location:</strong> {facility.full_address}</p>
                <p><strong>Phone:</strong> {facility.phone_number}</p>
            </div>
            
            <h3 style="color: #333;">Call Details</h3>
            <p><strong>Call Status:</strong> {call_session.get_call_status_display()}</p>
            <p><strong>Duration:</strong> {call_session.duration_seconds:.1f} seconds</p>
            <p><strong>Call Date:</strong> {call_session.created_at.strftime('%Y-%m-%d %H:%M:%S')}</p>
            
            {"<h3 style='color: #333;'>Call Summary</h3><p>" + summary.summary_text + "</p>" if summary else ""}
            
            {"<h3 style='color: #333;'>Key Points</h3><ul>" + "".join([f"<li>{point}</li>" for point in summary.key_points]) + "</ul>" if summary and summary.key_points else ""}
            
            <h3 style="color: #333;">Next Steps</h3>
            <p>Thank you for speaking with our team. We will be in touch with more information about our services.</p>
            
            <p style="color: #666; font-size: 12px; margin-top: 30px;">
                This is an automated message. Please do not reply to this email.
            </p>
        </div>
        """
        
        # Create email
        sg = sendgrid.SendGridAPIClient(api_key=settings.SENDGRID_API_KEY)
        message = Mail(
            from_email='marketing@pinksurfing.com',
            to_emails=settings.DEFAULT_FROM_EMAIL,  # Send to admin email, can be updated
            subject=f'Call Summary - {facility.facility_name}',
            html_content=email_html
        )
        
        # Add recording URL as attachment if available
        if call_session.recording_url:
            # For now, we'll skip adding the actual recording as attachment
            # In production, you might want to download and attach it
            pass
        
        # Send email
        sg.send(message)
        logger.info(f"Email sent successfully to facility: {facility.facility_name}")
        return True
    
    except Exception as e:
        logger.error(f"Failed to send email for facility {facility.facility_name}: {str(e)}")
        return False
