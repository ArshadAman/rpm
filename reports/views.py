from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from .models import Reports, Documentation
from rpm_users.models import Patient, Moderator
from .serializers import ReportSerializer
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from .forms import DocumentationForm  # Ensure you have a Django form for validation
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import date
from django.forms.models import model_to_dict
import re

# from rpm.customPermission import CustomSSOAuthentication

from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Excel styling constants
EXCEL_HEADER_COLOR = "7928CA"
EXCEL_HEADER_FONT_COLOR = "FFFFFF"
EXCEL_PATIENT_INFO_COLOR = "E0C3FC"

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_your_reports(request, patient_id):
    email = request.email
    if patient := Patient.objects.filter(email=email).first():
        try:
            reports = Reports.objects.filter(Q(patient=patient)).order_by("-created_at")
            serializer = ReportSerializer(reports, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Reports.DoesNotExist:
            return Response(
                {"error": "Reports not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
    return Response({"error": "Patient not exists"}, status=status.HTTP_404_NOT_FOUND)

@login_required
def get_patients_reports(request):
    email = request.user.email
    if moderator := Moderator.objects.filter(email=email).first():
        try:
            reports = Reports.objects.filter(
                patient__moderator_assigned=moderator
            ).order_by("-created_at")
            serializer = ReportSerializer(reports, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Reports.DoesNotExist:
            return Response(
                {"error": "Reports not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
    return Response({"error": "Moderator not exists"}, status=status.HTTP_404_NOT_FOUND)

@login_required
def get_single_report(request, report_id):
    user = request.user  # Get the logged-in user
    print(f"DEBUG: Logged-in User = {user}")

    try:
        report = Reports.objects.get(id=report_id)
        print(f"DEBUG: Report Found - ID: {report.id}")

        # Allow access if the user is a moderator or the assigned patient
        if Moderator.objects.filter(user=user).exists() or user == report.patient.user:
            print("DEBUG: Access granted")

            # Fetch related documentations
            documentations = report.documentations.all()

            context = {
                "report": report,
                "documentations": documentations,
            }

            return render(request, "index.html", context)

        print("DEBUG: Access denied - Only authorized users can view this report")
        return render(request, "errors/403.html", {"error": "You are not authorized to view this report"}, status=403)

    except Reports.DoesNotExist:
        print("DEBUG: Report not found")
        return render(request, "errors/404.html", {"error": "Report not found"}, status=404)


@login_required
def get_all_reports(request, patient_id):
    user = request.user  # Get the logged-in user
    print(f"DEBUG: Logged-in User = {user}")

    # Check if user is a Moderator
    is_moderator = Moderator.objects.filter(user=user).exists()

    # Fetch reports - If user is a moderator, fetch all reports, else fetch only the patient's reports
    patient = Patient.objects.filter(id=patient_id).first()
    print(f"DEBUG: Patient = {patient}")
    
    if not patient:
        print("DEBUG: Patient not found")
        return JsonResponse({"error": "Patient not found"}, status=404)
    
    # Check if the logged-in user is the patient or a moderator
    is_patient = Patient.objects.filter(user=user, id=patient_id).exists()
    
    if is_moderator or is_patient:
        reports = Reports.objects.filter(patient=patient).order_by("-created_at")
        print("reports", reports)
    else:
        print(f"DEBUG: User {user} is not authorized to view reports for patient {patient_id}")
        return JsonResponse({"error": "Not authorized to view these reports"}, status=403)

    # Return empty reports array if no reports exist
    if not reports.exists():
        print("DEBUG: No reports found, returning empty array")
        return JsonResponse({"reports": []}, safe=False)

    # Prepare JSON response with all fields
    data = {
        "reports": [
            {**model_to_dict(report), "created_at": report.created_at.strftime("%Y-%m-%d %H:%M:%S")} for report in reports
        ]
    }

    return JsonResponse(data, safe=False)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_report(request):
    email = request.email
    if patient := Patient.objects.filter(email=email).first():
        data = request.data

        # List all model fields you want to accept
        report_fields = {
            "blood_pressure": data.get("blood_pressure", ""),
            "heart_rate": data.get("heart_rate", ""),
            "spo2": data.get("spo2", ""),
            "temperature": data.get("temperature", ""),
            "symptoms": data.get("symptoms", ""),
            "device_id": data.get("device_id", ""),
            "created_at_device": data.get("created_at_device", ""),
            "data_type": data.get("data_type", ""),
            "imei": data.get("imei", ""),
            "iccid": data.get("iccid", ""),
            "serial_number": data.get("serial_number", ""),
            "model_number": data.get("model_number", ""),
            "is_test": data.get("is_test", ""),
            "user_id": data.get("user_id", ""),
            "systolic_blood_pressure": data.get("systolic_blood_pressure", ""),
            "diastolic_blood_pressure": data.get("diastolic_blood_pressure", ""),
            "pulse": data.get("pulse", ""),
            "irregular_heartbeat": data.get("irregular_heartbeat", ""),
            "hand_shaking": data.get("hand_shaking", ""),
            "triple_mode": data.get("triple_mode", ""),
            "battery_level": data.get("battery_level", ""),
            "signal_strength": data.get("signal_strength", ""),
            "measurement_timestamp": data.get("measurement_timestamp", ""),
            "timezone": data.get("timezone", ""),
            "blood_glucose": data.get("blood_glucose", ""),
            "glucose_unit": data.get("glucose_unit", ""),
            "test_paper_type": data.get("test_paper_type", ""),
            "sample_type": data.get("sample_type", ""),
            "meal_mark": data.get("meal_mark", ""),
            "signal_level": data.get("signal_level", ""),
            "measurement_timezone": data.get("measurement_timezone", ""),
            "upload_timestamp": data.get("upload_timestamp", ""),
            "upload_timezone": data.get("upload_timezone", ""),
        }

        report = Reports.objects.create(patient=patient, **report_fields)
        serializer = ReportSerializer(report)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response({"error": "Patient not exists"}, status=status.HTTP_404_NOT_FOUND)

@login_required
def add_documentation(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    
    # Get reports from the last month
    one_month_ago = timezone.now() - timedelta(days=30)
    reports = Reports.objects.filter(
        patient=patient,
    ).order_by('-created_at')
    # Calculate min/max values for vitals
    min_bp = None
    max_bp = None
    min_hr = None
    max_hr = None
    min_spo2 = None
    max_spo2 = None
    min_temp = None
    max_temp = None

    for report in reports:
        # Blood Pressure - try both old and new fields
        bp_value = None
        if report.blood_pressure:
            try:
                bp_values = [int(x.strip()) for x in report.blood_pressure.split('/')]
                if len(bp_values) == 2:
                    bp_value = bp_values[0]
            except (ValueError, AttributeError):
                pass
        elif report.systolic_blood_pressure:
            try:
                bp_value = int(report.systolic_blood_pressure)
            except (ValueError, AttributeError):
                pass

        if bp_value is not None:
            if min_bp is None or bp_value < min_bp:
                min_bp = bp_value
            if max_bp is None or bp_value > max_bp:
                max_bp = bp_value

        # Heart Rate - try both old and new fields
        hr_value = None
        if report.heart_rate:
            try:
                hr_value = int(report.heart_rate)
            except (ValueError, AttributeError):
                pass
        elif report.pulse:
            try:
                hr_value = int(report.pulse)
            except (ValueError, AttributeError):
                pass

        if hr_value is not None:
            if min_hr is None or hr_value < min_hr:
                min_hr = hr_value
            if max_hr is None or hr_value > max_hr:
                max_hr = hr_value

        # SpO2
        if report.spo2:
            try:
                spo2 = int(report.spo2)
                if min_spo2 is None or spo2 < min_spo2:
                    min_spo2 = spo2
                if max_spo2 is None or spo2 > max_spo2:
                    max_spo2 = spo2
            except (ValueError, AttributeError):
                pass

        # Temperature
        if report.temperature:
            try:
                temp = float(report.temperature)
                if min_temp is None or temp < min_temp:
                    min_temp = temp
                if max_temp is None or temp > max_temp:
                    max_temp = temp
            except (ValueError, AttributeError):
                pass
    print("min_bp", min_bp)
    print("max_bp", max_bp)
    print("min_hr", min_hr)
    print("max_hr", max_hr)
    print("min_spo2", min_spo2)
    print("max_spo2", max_spo2)
    print("min_temp", min_temp)
    if request.method == 'POST':
        form = DocumentationForm(request.POST, request.FILES)
        if form.is_valid():
            documentation = form.save(commit=False)
            documentation.patient = patient
            
            # Update patient snapshot fields
            documentation.doc_patient_name = f"{patient.user.first_name} {patient.user.last_name}"
            documentation.doc_dob = patient.date_of_birth
            documentation.doc_sex = patient.get_sex_display()
            documentation.doc_monitoring_params = patient.monitoring_parameters
            documentation.doc_clinical_staff = str(patient.moderator_assigned) if patient.moderator_assigned else "N/A"
            documentation.doc_moderator = str(patient.moderator_assigned) if patient.moderator_assigned else "N/A"
            documentation.doc_report_date = timezone.now().date()
            # Update the history_of_present_illness with the full_documentation value
            documentation.history_of_present_illness = request.POST.get('full_documentation', '')
            documentation.written_by = request.user.username
            documentation.doc_report_date = request.POST.get('doc_report_date', '')
            documentation.save()
            messages.success(request, 'Documentation added successfully.')
            return redirect('moderator_actions', patient_id=str(patient.id))
    else:
        form = DocumentationForm()

    context = {
        'form': form,
        'patient': patient,
        'reports': reports,
        'min_bp': min_bp,
        'max_bp': max_bp,
        'min_hr': min_hr,
        'max_hr': max_hr,
        'min_spo2': min_spo2,
        'max_spo2': max_spo2,
        'min_temp': min_temp,
        'max_temp': max_temp,
        'now': timezone.now()
    }
    print("context", context)
    return render(request, 'reports/add_docs.html', context)


@csrf_exempt
def data_from_mio_connect(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=405)

    try:
        # Parse JSON data
        try:
            body = json.loads(request.body)
            print(f"Received data from MioConnect: {json.dumps(body)}")
        except json.JSONDecodeError as e:
            return JsonResponse({"error": f"Invalid JSON data: {str(e)}"}, status=400)

        data = body.get('data', {})
        device_serial = data.get('sn', '')
        if not device_serial:
            return JsonResponse({"error": "Missing device serial number"}, status=400)

        patient = Patient.objects.filter(device_serial_number=device_serial).first()
        if not patient:
            return JsonResponse({"error": f"No patient found with device serial number: {device_serial}"}, status=404)

        # List all possible fields from both device types
        report_fields = {
            # Common
            'device_id': body.get('deviceId', ''),
            'created_at_device': str(body.get('createdAt', '')),
            'data_type': data.get('data_type', ''),
            'imei': data.get('imei', ''),
            'iccid': data.get('iccid', ''),
            'serial_number': data.get('sn', ''),
            'model_number': body.get('modelNumber', ''),
            'is_test': str(body.get('isTest', '')),
            # Sphygmomanometer
            'user_id': str(data.get('user', '')),
            'systolic_blood_pressure': str(data.get('sys', '')),
            'diastolic_blood_pressure': str(data.get('dia', '')),
            'pulse': str(data.get('pul', '')),
            'irregular_heartbeat': str(data.get('ihb', '')),
            'hand_shaking': str(data.get('hand', '')),
            'triple_mode': str(data.get('tri', '')),
            'battery_level': str(data.get('bat', '')),
            'signal_strength': str(data.get('sig', '')),
            'measurement_timestamp': str(data.get('ts', '')),
            'timezone': str(data.get('tz', '')),
            # Blood Glucose Meter
            'blood_glucose': str(data.get('data', '')),
            'glucose_unit': str(data.get('unit', '')),
            'test_paper_type': str(data.get('sample', '')),
            'sample_type': str(data.get('target', '')),
            'meal_mark': str(data.get('meal', '')),
            'signal_level': str(data.get('sig_lvl', '')),
            'measurement_timezone': str(data.get('ts_tz', '')),
            'upload_timestamp': str(data.get('uptime', '')),
            'upload_timezone': str(data.get('uptime_tz', '')),
        }

        # For compatibility, also fill old fields if possible
        if report_fields['systolic_blood_pressure'] and report_fields['diastolic_blood_pressure']:
            report_fields['blood_pressure'] = f"{report_fields['systolic_blood_pressure']}/{report_fields['diastolic_blood_pressure']}"
        else:
            report_fields['blood_pressure'] = ''
        report_fields['heart_rate'] = report_fields['pulse']
        report_fields['spo2'] = ''
        report_fields['temperature'] = ''
        report_fields['symptoms'] = ''
        try:
            report = Reports.objects.create(patient=patient, **report_fields)
            # Send the email using sendgrid
            # Check the critical fields
            if int(report_fields['systolic_blood_pressure']) > 170:
                pass #send the email

        except Exception as e:
            print(f"Error creating report: {str(e)}")
            return JsonResponse({"error": f"Failed to create report: {str(e)}"}, status=500)
        return JsonResponse({
            "success": True,
            "received_data": body,
            "saved_report_id": report.id,
            "patient_id": patient.id
        })
    except Exception as err:
        print(f"Unexpected error: {str(err)}")
        return JsonResponse({"error": "An unexpected error occurred while processing the request"}, status=500)


@login_required
def view_documentation(request):
    # Check if the user is a moderator
    if not Moderator.objects.filter(user=request.user).exists():
        return render(request, "errors/403.html", {"error": "Only moderators can view documentation"}, status=403)
    
    # Get the date filter from the request
    date = request.GET.get('date')
    
    # Get all documentation for the moderator's patients
    moderator = Moderator.objects.get(user=request.user)
    patient_ids = Patient.objects.filter(moderator_assigned=moderator).values_list("id", flat=True)
    
    # Apply date filter if provided
    if date:
        documentations = Documentation.objects.filter(
            patient__in=patient_ids,
            created_at__date=date
        ).select_related('patient').order_by('-created_at')
    else:
        documentations = Documentation.objects.filter(
            patient__in=patient_ids
        ).select_related('patient').order_by('-created_at')
    
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if is_ajax:
        # For AJAX requests, return only the documentation cards
        return render(request, "reports/documentation_cards.html", {
            'documentations': documentations
        })
    else:
        # For regular requests, return the full page
        return render(request, "reports/view_documentation.html", {
            'documentations': documentations
        })

@login_required
def edit_patient(request, patient_id):
    user = request.user
    
    # Check if the user is a moderator or the patient themselves
    is_moderator = Moderator.objects.filter(user=user).exists()
    is_patient = Patient.objects.filter(user=user, id=patient_id).exists()
    
    if not is_moderator and not is_patient:
        return render(request, "errors/403.html", {"error": "You are not authorized to edit this patient information"}, status=403)
    
    patient = get_object_or_404(Patient, id=patient_id)
    
    if request.method == "POST":
        try:
            # Update user information with defaults
            patient.user.first_name = request.POST.get('first_name', '') or ''
            patient.user.last_name = request.POST.get('last_name', '') or ''
            # Admin-only: update patient account email/password if provided
            if user.is_superuser:
                new_email = (request.POST.get('user_email') or '').strip()
                if new_email:
                    patient.user.email = new_email
                    patient.user.username = new_email  # Assuming username is email
                new_password = (request.POST.get('user_password') or '').strip()
                if new_password:
                    patient.user.set_password(new_password)
            patient.user.save()

            # Update patient information with safe defaults
            from datetime import datetime
            date_of_birth_str = request.POST.get('date_of_birth', '')
            if date_of_birth_str:
                try:
                    patient.date_of_birth = datetime.strptime(date_of_birth_str, '%Y-%m-%d').date()
                except Exception:
                    patient.date_of_birth = None
            else:
                patient.date_of_birth = None

            patient.sex = request.POST.get('sex', patient.sex or 'Others') or 'Others'

            try:
                patient.weight = float(request.POST.get('weight', patient.weight or 0.0) or 0.0)
            except Exception:
                patient.weight = 0.0

            try:
                patient.height = float(request.POST.get('height', patient.height or 0.0) or 0.0)
            except Exception:
                patient.height = 0.0

            patient.insurance = request.POST.get('insurance', patient.insurance or '') or ''
            patient.insurance_number = request.POST.get('insurance_number', patient.insurance_number or '') or ''
            patient.monitoring_parameters = request.POST.get('monitoring_parameters', patient.monitoring_parameters or '') or ''

            device_serial_number = request.POST.get('device_serial_number', '')
            try:
                patient.device_serial_number = int(device_serial_number) if device_serial_number else None
            except Exception:
                patient.device_serial_number = None

            patient.drink = request.POST.get('drink', patient.drink or 'NO') or 'NO'
            patient.smoke = request.POST.get('smoke', patient.smoke or 'NO') or 'NO'
            patient.phone_number = request.POST.get('phone_number', patient.phone_number or '') or ''
            patient.allergies = request.POST.get('allergies', patient.allergies or '') or ''
            patient.family_history = request.POST.get('family_history', patient.family_history or '') or ''
            patient.pharmacy_info = request.POST.get('pharmacy_info', patient.pharmacy_info or '') or ''
            patient.medications = request.POST.get('medications', patient.medications or '') or ''
            patient.home_address = request.POST.get('home_address', patient.home_address or '') or ''
            patient.emergency_contact_name = request.POST.get('emergency_contact_name', patient.emergency_contact_name or '') or ''
            patient.emergency_contact_phone = request.POST.get('emergency_contact_phone', patient.emergency_contact_phone or '') or ''
            patient.emergency_contact_relationship = request.POST.get('emergency_contact_relationship', patient.emergency_contact_relationship or '') or ''
            patient.primary_care_physician = request.POST.get('primary_care_physician', patient.primary_care_physician or '') or ''
            patient.primary_care_physician_phone = request.POST.get('primary_care_physician_phone', patient.primary_care_physician_phone or '') or ''

            # Handle Past Medical History
            selected_pmh = request.POST.getlist('past_medical_history', [])
            # Delete existing past medical history
            patient.medical_history.all().delete()
            from rpm_users.models import PastMedicalHistory
            if selected_pmh:
                for pmh in selected_pmh:
                    PastMedicalHistory.objects.create(
                        patient=patient,
                        pmh=pmh
                    )
            else:
                # If nothing selected, add N/A
                PastMedicalHistory.objects.create(
                    patient=patient,
                    pmh='N/A'
                )

            # Recalculate BMI
            if patient.weight and patient.height:
                try:
                    height_in_meters = float(patient.height) / 100
                    if height_in_meters > 0:
                        patient.bmi = round(float(patient.weight) / (height_in_meters ** 2), 2)
                    else:
                        patient.bmi = 0.0
                except Exception:
                    patient.bmi = 0.0
            else:
                patient.bmi = 0.0

            patient.save()

            return JsonResponse({
                'success': True,
                'patient_id': patient.id
            })

        except Exception as e:
            # Log the error if needed, but always return a default error
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # GET request - display the edit form
    from rpm_users.models import PastMedicalHistory
    context = {
        'patient': patient,
        'pmh_choices': PastMedicalHistory.PMH_CHOICES,
        'is_admin': request.user.is_superuser
    }
    return render(request, "reports/edit_patient.html", context)

@login_required
def edit_documentation(request, doc_id):
    documentation = get_object_or_404(Documentation, id=doc_id)
    patient = documentation.patient
    print("documentation", documentation,patient)
    # Check if the user is a moderator
    if not Moderator.objects.filter(user=request.user).exists():
        return render(request, "errors/403.html", {"error": "Only moderators can edit documentation"}, status=403)
    
    if request.method == 'POST':
        form = DocumentationForm(request.POST, request.FILES, instance=documentation)
        if form.is_valid():
            documentation = form.save(commit=False)
            # Update patient snapshot fields
            documentation.doc_patient_name = request.POST.get('doc_patient_name', f"{patient.user.first_name} {patient.user.last_name}")
            documentation.doc_dob = request.POST.get('doc_dob', patient.date_of_birth)
            documentation.doc_sex = request.POST.get('doc_sex', patient.get_sex_display())
            documentation.doc_monitoring_params = request.POST.get('doc_monitoring_params', patient.monitoring_parameters)
            documentation.doc_clinical_staff = request.POST.get('doc_clinical_staff', str(patient.moderator_assigned) if patient.moderator_assigned else "N/A")
            documentation.doc_moderator = request.POST.get('doc_moderator', str(patient.moderator_assigned) if patient.moderator_assigned else "N/A")
            # Date of Service editable field
            doc_report_date = request.POST.get('doc_report_date')
            if doc_report_date:
                documentation.doc_report_date = doc_report_date
            else:
                documentation.doc_report_date = timezone.now().date()
            # Update the history_of_present_illness with the full_documentation value
            documentation.history_of_present_illness = request.POST.get('full_documentation', '')
            documentation.save()
            messages.success(request, 'Documentation updated successfully.')
            return redirect('moderator_actions', patient_id=str(patient.id))
    else:
        # Pre-fill the form with existing documentation data
        initial_data = {
            'title': documentation.title,
            'doc_patient_name': documentation.doc_patient_name or f"{patient.user.first_name} {patient.user.last_name}",
            'doc_dob': documentation.doc_dob or patient.date_of_birth,
            'doc_sex': documentation.doc_sex or patient.get_sex_display(),
            'doc_monitoring_params': documentation.doc_monitoring_params or patient.monitoring_parameters,
            'doc_clinical_staff': documentation.doc_clinical_staff or (str(patient.moderator_assigned) if patient.moderator_assigned else "N/A"),
            'doc_moderator': documentation.doc_moderator or (str(patient.moderator_assigned) if patient.moderator_assigned else "N/A"),
            # Date of Service field
            'doc_report_date': documentation.doc_report_date or timezone.now().date(),
            'full_documentation': documentation.history_of_present_illness
        }
        print("initial_data", initial_data)
        form = DocumentationForm(instance=documentation, initial=initial_data)
        print("Form",form)
    context = {
        'form': form,
        'documentation': documentation,
        'doc_id': documentation.id,  # <-- Add this line
        'patient': patient,
        # 'reports': reports,
        'now': timezone.now(),
        'full_documentation': documentation.history_of_present_illness
    }
    
    return render(request, 'reports/edit_docs.html', context)

def documentation_share_view(request, doc_id):
    doc = get_object_or_404(Documentation, id=doc_id)
    return render(request, 'documentation_share.html', {'doc': doc, 'now': timezone.now()})
# Ne

@login_required
@csrf_exempt
def update_report(request, report_id):
    """Update an existing report - Moderator only"""
    if request.method != 'PUT':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    # Check if user is a moderator
    if not Moderator.objects.filter(user=request.user).exists():
        return JsonResponse({'success': False, 'error': 'Only moderators can update reports'}, status=403)
    
    try:
        # Get the report
        report = get_object_or_404(Reports, id=report_id)
        
        # Parse JSON data
        data = json.loads(request.body)
        
        # Validate and update fields
        valid_fields = [
            'systolic_blood_pressure', 'diastolic_blood_pressure', 'pulse', 'heart_rate',
            'spo2', 'temperature', 'blood_glucose', 'glucose_unit', 'symptoms',
            'blood_pressure', 'irregular_heartbeat', 'hand_shaking', 'battery_level'
        ]
        
        errors = {}
        
        # Validate numeric fields
        numeric_fields = {
            'systolic_blood_pressure': (70, 250),
            'diastolic_blood_pressure': (40, 150),
            'pulse': (30, 200),
            'heart_rate': (30, 200),
            'spo2': (70, 100),
            'temperature': (95.0, 110.0),
            'blood_glucose': (50, 500)
        }
        
        for field, value in data.items():
            if field in valid_fields:
                if field in numeric_fields and value:
                    try:
                        num_value = float(value)
                        min_val, max_val = numeric_fields[field]
                        if not (min_val <= num_value <= max_val):
                            errors[field] = f'Value must be between {min_val} and {max_val}'
                    except ValueError:
                        errors[field] = 'Must be a valid number'
                
                # Update the field if no errors
                if field not in errors:
                    setattr(report, field, value)
        
        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        
        # Save the updated report
        report.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Report updated successfully',
            'report': model_to_dict(report)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@csrf_exempt
def create_report_manual(request):
    """Create a new report manually - Moderator only"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    # Check if user is a moderator
    if not Moderator.objects.filter(user=request.user).exists():
        return JsonResponse({'success': False, 'error': 'Only moderators can create reports'}, status=403)
    
    try:
        # Parse JSON data
        data = json.loads(request.body)
        
        # Get patient
        patient_id = data.get('patient_id')
        if not patient_id:
            return JsonResponse({'success': False, 'error': 'Patient ID is required'}, status=400)
        
        patient = get_object_or_404(Patient, id=patient_id)
        
        # Validate numeric fields
        numeric_fields = {
            'systolic_blood_pressure': (70, 250),
            'diastolic_blood_pressure': (40, 150),
            'pulse': (30, 200),
            'heart_rate': (30, 200),
            'spo2': (70, 100),
            'temperature': (95.0, 110.0),
            'blood_glucose': (50, 500)
        }
        
        errors = {}
        report_fields = {}
        
        # Process and validate fields
        for field, value in data.items():
            if field == 'patient_id':
                continue
                
            if field in numeric_fields and value:
                try:
                    num_value = float(value)
                    min_val, max_val = numeric_fields[field]
                    if not (min_val <= num_value <= max_val):
                        errors[field] = f'Value must be between {min_val} and {max_val}'
                    else:
                        # Save as int if value is whole number, else as float string
                        if field in ['systolic_blood_pressure', 'diastolic_blood_pressure', 'pulse', 'heart_rate', 'spo2', 'blood_glucose']:
                            # These fields should be saved as integer strings
                            report_fields[field] = str(int(num_value))
                        else:
                            # For temperature, save as float string
                            report_fields[field] = str(num_value)
                except ValueError:
                    errors[field] = 'Must be a valid number'
            elif value:  # Non-numeric fields
                report_fields[field] = str(value)
        
        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        
        # Create combined blood pressure field if both systolic and diastolic are provided
        if 'systolic_blood_pressure' in report_fields and 'diastolic_blood_pressure' in report_fields:
            report_fields['blood_pressure'] = f"{report_fields['systolic_blood_pressure']}/{report_fields['diastolic_blood_pressure']}"
        
        # Set default values for required fields
        report_fields.update({
            'measurement_timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'created_at_device': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'data_type': 'manual_entry',
            'is_test': 'false'
        })
        
        # Create the report
        report = Reports.objects.create(patient=patient, **report_fields)
        
        return JsonResponse({
            'success': True,
            'message': 'Report created successfully',
            'report': model_to_dict(report)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def get_recent_reports(request, patient_id):
    """Get recent reports for a patient - for edit dropdown"""
    # Check if user is a moderator
    if not Moderator.objects.filter(user=request.user).exists():
        return JsonResponse({'success': False, 'error': 'Only moderators can access this endpoint'}, status=403)
    
    try:
        patient = get_object_or_404(Patient, id=patient_id)
        
        # Get reports from last 30 days
        thirty_days_ago = timezone.now() - timedelta(days=30)
        reports = Reports.objects.filter(
            patient=patient,
            created_at__gte=thirty_days_ago
        ).order_by('-created_at')[:20]  # Limit to 20 most recent
        
        reports_data = []
        for report in reports:
            # Create a summary of vital signs for display
            vitals_summary = []
            if report.systolic_blood_pressure and report.diastolic_blood_pressure:
                vitals_summary.append(f"BP: {report.systolic_blood_pressure}/{report.diastolic_blood_pressure}")
            elif report.blood_pressure:
                vitals_summary.append(f"BP: {report.blood_pressure}")
            
            if report.pulse:
                vitals_summary.append(f"HR: {report.pulse}")
            elif report.heart_rate:
                vitals_summary.append(f"HR: {report.heart_rate}")
            
            if report.spo2:
                vitals_summary.append(f"SpO2: {report.spo2}%")
            
            if report.temperature:
                vitals_summary.append(f"Temp: {report.temperature}°F")
            
            reports_data.append({
                'id': report.id,
                'created_at': report.created_at.strftime('%m/%d/%Y %H:%M'),
                'vitals_summary': ', '.join(vitals_summary) if vitals_summary else 'No vitals recorded',
                'data': model_to_dict(report)
            })
        
        return JsonResponse({
            'success': True,
            'reports': reports_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def export_vitals_excel(request, patient_id):
    """
    Export patient vitals data to Excel format with proper formatting
    """
    user = request.user
    
    # Check if user is a moderator or the patient themselves or superuser
    is_moderator = Moderator.objects.filter(user=user).exists()
    is_patient = Patient.objects.filter(user=user, id=patient_id).exists()
    
    if not is_moderator and not is_patient and not user.is_superuser:
        return JsonResponse({"error": "Not authorized to export this patient's data"}, status=403)
    
    try:
        patient = get_object_or_404(Patient, id=patient_id)
        reports = Reports.objects.filter(patient=patient).order_by('-created_at')
        
        if not reports.exists():
            return JsonResponse({"error": "No vitals data available to export"}, status=404)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Vitals Data"
        
        # Define styles using constants
        header_fill = PatternFill(start_color=EXCEL_HEADER_COLOR, end_color=EXCEL_HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color=EXCEL_HEADER_FONT_COLOR, size=12)
        patient_info_fill = PatternFill(start_color=EXCEL_PATIENT_INFO_COLOR, end_color=EXCEL_PATIENT_INFO_COLOR, fill_type="solid")
        patient_info_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Patient Information Section
        ws.merge_cells('A1:F1')
        ws['A1'] = 'PATIENT VITALS REPORT'
        ws['A1'].font = Font(bold=True, size=16, color="7928CA")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A3'] = 'Patient Name:'
        ws['A3'].font = patient_info_font
        ws['B3'] = f"{patient.user.first_name} {patient.user.last_name}"
        
        ws['A4'] = 'Date of Birth:'
        ws['A4'].font = patient_info_font
        ws['B4'] = str(patient.date_of_birth) if patient.date_of_birth else 'N/A'
        
        ws['A5'] = 'Email:'
        ws['A5'].font = patient_info_font
        ws['B5'] = patient.user.email
        
        ws['A6'] = 'Phone:'
        ws['A6'].font = patient_info_font
        ws['B6'] = patient.phone_number if patient.phone_number else 'N/A'
        
        ws['A7'] = 'Report Generated:'
        ws['A7'].font = patient_info_font
        ws['B7'] = timezone.now().strftime('%m/%d/%Y %H:%M')
        
        # Headers for vitals data (starting at row 9)
        headers = [
            'Date/Time',
            'Systolic BP (mmHg)',
            'Diastolic BP (mmHg)',
            'Heart Rate (bpm)',
            'SpO2 (%)',
            'Temperature (°F)',
            'Blood Glucose',
            'Irregular Heartbeat',
            'Symptoms'
        ]
        
        header_row = 9
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row_num = header_row + 1
        for report in reports:
            # Date/Time
            ws.cell(row=row_num, column=1).value = report.created_at.strftime('%m/%d/%Y %H:%M')
            
            # Systolic BP - with safe parsing
            systolic = report.systolic_blood_pressure or ''
            if not systolic and report.blood_pressure and '/' in report.blood_pressure:
                try:
                    bp_parts = report.blood_pressure.split('/')
                    if len(bp_parts) >= 2:
                        systolic = bp_parts[0].strip()
                except (AttributeError, IndexError):
                    systolic = ''
            ws.cell(row=row_num, column=2).value = systolic
            
            # Diastolic BP - with safe parsing
            diastolic = report.diastolic_blood_pressure or ''
            if not diastolic and report.blood_pressure and '/' in report.blood_pressure:
                try:
                    bp_parts = report.blood_pressure.split('/')
                    if len(bp_parts) >= 2:
                        diastolic = bp_parts[1].strip()
                except (AttributeError, IndexError):
                    diastolic = ''
            ws.cell(row=row_num, column=3).value = diastolic
            
            # Heart Rate
            heart_rate = report.pulse or report.heart_rate or ''
            ws.cell(row=row_num, column=4).value = heart_rate
            
            # SpO2
            ws.cell(row=row_num, column=5).value = report.spo2 or ''
            
            # Temperature
            ws.cell(row=row_num, column=6).value = report.temperature or ''
            
            # Blood Glucose
            ws.cell(row=row_num, column=7).value = report.blood_glucose or ''
            
            # Irregular Heartbeat
            ws.cell(row=row_num, column=8).value = report.irregular_heartbeat or ''
            
            # Symptoms
            ws.cell(row=row_num, column=9).value = report.symptoms or ''
            
            # Apply borders and alignment to data cells
            for col_num in range(1, 10):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 30
        
        # Freeze header row
        ws.freeze_panes = ws['A10']
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Sanitize filename to prevent path traversal and special character issues
        safe_last_name = re.sub(r'[^\w\s-]', '', patient.user.last_name)[:50]
        safe_first_name = re.sub(r'[^\w\s-]', '', patient.user.first_name)[:50]
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"vitals_{safe_last_name}_{safe_first_name}_{timestamp}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error exporting vitals: {str(e)}")
        return JsonResponse({"error": f"Failed to export vitals: {str(e)}"}, status=500)
