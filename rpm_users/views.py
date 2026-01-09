from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
import requests
import sendgrid
from sendgrid.helpers.mail import Mail
from django.conf import settings

from reports.models import Reports, Documentation
from reports.serializers import ReportSerializer
from reports.forms import ReportForm
from .models import Patient, Moderator, PastMedicalHistory, Interest, InterestPastMedicalHistory, InterestLead, Doctor, EmailOTP, LabCategory, LabTest, LabResult, LabDocument
from retell_calling.models import CallSummary, LeadCallSession, LeadCallSummary
from referral.models import Referral
from django.db import models
from .serializers import PatientSerializer, ModeratorSerializer
from django.contrib.auth.hashers import make_password
from django.contrib.auth import get_user_model
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, Http404
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from .form import DocumentationForm
from .forms import PatientForm
from django.db import transaction

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.sites import site
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import transaction
import re
import logging
from datetime import datetime
from django_ratelimit.decorators import ratelimit
import random
from datetime import timedelta

def send_otp_email(email, otp_code):
    """Send OTP verification email using SendGrid"""
    try:
        # Skip email sending if no SendGrid API key is configured
        if not hasattr(settings, 'SENDGRID_API_KEY') or not settings.SENDGRID_API_KEY:
            logging.warning(f"SendGrid not configured. Skipping OTP email for {email}")
            return False
        
        # Use SendGrid to send email
        sg = sendgrid.SendGridAPIClient(api_key=settings.SENDGRID_API_KEY)
        message = Mail(
            from_email='marketing@pinksurfing.com',
            to_emails=email,
            subject='RPM Patient Registration - Email Verification',
            html_content=f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #7928CA; text-align: center;">Email Verification</h2>
                <p>Thank you for registering with RPM Patient Portal. Please use the following verification code to complete your registration:</p>
                
                <div style="background: linear-gradient(135deg, #7928CA, #FF0080); color: white; padding: 20px; text-align: center; border-radius: 10px; margin: 20px 0;">
                    <h1 style="margin: 0; font-size: 32px; letter-spacing: 5px;">{otp_code}</h1>
                </div>
                
                <p><strong>Important:</strong></p>
                <ul>
                    <li>This code will expire in 10 minutes</li>
                    <li>Do not share this code with anyone</li>
                    <li>If you didn't request this verification, please ignore this email</li>
                </ul>
                
                <p style="color: #666; font-size: 14px; margin-top: 30px;">
                    If you have any questions, please contact our support team.
                </p>
            </div>
            """
        )
        
        # Send email
        sg.send(message)
        logging.info(f"OTP email sent successfully to {email}")
        return True
        
    except Exception as e:
        logging.error(f"Failed to send OTP email to {email}: {str(e)}")
        return False

def generate_otp():
    """Generate a 6-digit OTP code"""
    return str(random.randint(100000, 999999))

def home(request):
    """Homepage with options for moderator login and patient registration"""
    return render(request, 'home.html')

def landing_page(request):
    """Public landing page mapped to '/' route"""
    return render(request, 'landing_page.html')

def admin_login(request):
    """Custom admin login page that only allows superusers"""
    if request.user.is_authenticated and request.user.is_superuser:
        # User is already authenticated as admin, redirect to dashboard
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_superuser:
                login(request, user)
                messages.success(request, 'Successfully logged in as administrator.')
                return redirect('admin_dashboard')
            else:
                messages.error(request, 'Access denied. Administrator privileges required.')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'admin_login.html')

def admin_access(request):
    """Check if user is admin and redirect appropriately"""
    if request.user.is_authenticated and request.user.is_superuser:
        # User is authenticated and is a superuser, redirect to admin dashboard
        return redirect('admin_dashboard')
    else:
        # User is not authenticated or not a superuser, redirect to custom admin login
        return redirect('admin_login')

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_dashboard(request):
    """Admin dashboard with three main action sections"""
    try:
        from .models import Video
        
        # Get counts for dashboard display
        moderator_count = Moderator.objects.count()
        doctor_count = Doctor.objects.count()
        patient_count = Patient.objects.count()
        call_summary_count = CallSummary.objects.count()
        
        # Get leads statistics
        total_leads = InterestLead.objects.count()
        converted_leads = InterestLead.objects.filter(is_converted=True).count()
        conversion_rate = round((converted_leads / total_leads * 100), 1) if total_leads > 0 else 0
        
        # Get lead call statistics
        lead_calls_count = LeadCallSession.objects.count()
        leads_with_calls_count = InterestLead.objects.annotate(
            call_count=models.Count('lead_call_summaries')
        ).filter(call_count__gt=0).count()
        
        # Get referral statistics
        total_referrals = Referral.objects.count()
        pending_referrals = Referral.objects.filter(contacted=False).count()
        
        # Get video statistics
        video_count = Video.objects.count()
        active_videos = Video.objects.filter(is_active=True).count()
        
        # Get testimonial statistics (with fallback if table doesn't exist)
        try:
            from .models import Testimonial
            testimonial_count = Testimonial.objects.count()
            active_testimonials = Testimonial.objects.filter(is_active=True).count()
        except Exception:
            testimonial_count = 0
            active_testimonials = 0
        
        context = {
            'moderator_count': moderator_count,
            'doctor_count': doctor_count,
            'patient_count': patient_count,
            'call_summary_count': call_summary_count,
            'total_leads': total_leads,
            'converted_leads': converted_leads,
            'conversion_rate': conversion_rate,
            'lead_calls_count': lead_calls_count,
            'leads_with_calls_count': leads_with_calls_count,
            'total_referrals': total_referrals,
            'pending_referrals': pending_referrals,
            'video_count': video_count,
            'active_videos': active_videos,
            'testimonial_count': testimonial_count,
            'active_testimonials': active_testimonials,
        }
        
        return render(request, 'admin_dashboard.html', context)
    except Exception as e:
        messages.error(request, f'Error loading admin dashboard: {str(e)}')
        return redirect('home')

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def moderator_list(request):
    """Display list of all moderators with search and filtering capabilities"""
    try:
        # Get search query from request
        search_query = request.GET.get('search', '').strip()
        
        # Start with all moderators
        moderators = Moderator.objects.select_related('user').all()
        
        # Apply search filter if provided
        if search_query:
            moderators = moderators.filter(
                models.Q(user__first_name__icontains=search_query) |
                models.Q(user__last_name__icontains=search_query) |
                models.Q(user__email__icontains=search_query) |
                models.Q(user__username__icontains=search_query) |
                models.Q(phone_number__icontains=search_query)
            )
        
        # Order by username for consistent display
        moderators = moderators.order_by('user__username')
        
        # Add patient count for each moderator
        moderators_with_counts = []
        for moderator in moderators:
            patient_count = Patient.objects.filter(moderator_assigned=moderator).count()
            moderators_with_counts.append({
                'moderator': moderator,
                'patient_count': patient_count
            })
        
        context = {
            'moderators_with_counts': moderators_with_counts,
            'search_query': search_query,
            'total_count': len(moderators_with_counts)
        }
        
        return render(request, 'admin/moderator_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading moderator list: {str(e)}', extra_tags="admin")
        return redirect('admin_dashboard')

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def moderator_create(request):
    """Create a new moderator with form handling and validation"""
    if request.method == 'POST':
        from .forms import ModeratorForm
        form = ModeratorForm(request.POST)
        
        if form.is_valid():
            try:
                # Create the User object
                user = User.objects.create(
                    username=form.cleaned_data['username'],
                    email=form.cleaned_data['email'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name']
                )
                user.set_password(form.cleaned_data['password'])
                user.save()
                
                # Create the Moderator object
                moderator = Moderator.objects.create(
                    user=user,
                    phone_number=form.cleaned_data['phone_number']
                )
                
                messages.success(request, f'Moderator "{user.get_full_name()}" created successfully.', extra_tags="admin")
                return redirect('moderator_list')
                
            except Exception as e:
                messages.error(request, f'Error creating moderator: {str(e)}', extra_tags="admin")
                # Clean up user if moderator creation failed
                if 'user' in locals():
                    user.delete()
        else:
            messages.error(request, 'Please correct the errors below.', extra_tags="admin")
    else:
        from .forms import ModeratorForm
        form = ModeratorForm()
    
    context = {
        'form': form,
        'title': 'Create New Moderator',
        'action': 'Create'
    }
    
    return render(request, 'admin/moderator_form.html', context)

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def moderator_detail(request, moderator_id):
    """Display detailed information about a specific moderator including assigned patients"""
    try:
        moderator = get_object_or_404(Moderator, id=moderator_id)
        
        # Get all patients assigned to this moderator
        assigned_patients = Patient.objects.filter(moderator_assigned=moderator).select_related('user').order_by('user__first_name')
        
        # Add additional info for each patient
        patients_with_info = []
        for patient in assigned_patients:
            # Calculate age if date_of_birth exists
            age = None
            if patient.date_of_birth:
                from datetime import date
                today = date.today()
                age = today.year - patient.date_of_birth.year - ((today.month, today.day) < (patient.date_of_birth.month, patient.date_of_birth.day))
            
            # Get latest report for this patient (if any)
            from reports.models import Reports
            latest_report = Reports.objects.filter(patient=patient).order_by('-created_at').first()
            
            # Get escalated doctor information if patient is escalated
            escalated_doctor = None
            if patient.is_escalated and patient.doctor_escalated:
                escalated_doctor = patient.doctor_escalated
            
            patients_with_info.append({
                'patient': patient,
                'age': age,
                'latest_report': latest_report,
                'escalated': patient.is_escalated,
                'escalated_doctor': escalated_doctor
            })
        
        context = {
            'moderator': moderator,
            'patients_with_info': patients_with_info,
            'total_patients': len(patients_with_info),
            'escalated_count': len([p for p in patients_with_info if p['escalated']]),
            'normal_count': len([p for p in patients_with_info if not p['escalated']])
        }
        
        return render(request, 'admin/moderator_detail.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading moderator details: {str(e)}', extra_tags="admin")
        return redirect('moderator_list')

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def moderator_edit(request, moderator_id):
    """Edit an existing moderator with form handling and validation"""
    moderator = get_object_or_404(Moderator, id=moderator_id)
    
    if request.method == 'POST':
        from .forms import ModeratorForm
        form = ModeratorForm(request.POST, instance=moderator)
        
        if form.is_valid():
            try:
                # Update the User object
                user = moderator.user
                user.username = form.cleaned_data['username']
                user.email = form.cleaned_data['email']
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                
                # Update password if provided
                if form.cleaned_data['password']:
                    user.set_password(form.cleaned_data['password'])
                
                user.save()
                
                # Update the Moderator object
                moderator.phone_number = form.cleaned_data['phone_number']
                moderator.save()
                
                messages.success(request, f'Moderator "{user.get_full_name()}" updated successfully.', extra_tags="admin")
                return redirect('moderator_list')
                
            except Exception as e:
                messages.error(request, f'Error updating moderator: {str(e)}', extra_tags="admin")
        else:
            messages.error(request, 'Please correct the errors below.', extra_tags="admin")
    else:
        from .forms import ModeratorForm
        form = ModeratorForm(instance=moderator)
    
    context = {
        'form': form,
        'title': f'Edit Moderator: {moderator.user.get_full_name()}',
        'action': 'Update',
        'moderator': moderator
    }
    
    return render(request, 'admin/moderator_form.html', context)

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def moderator_delete(request, moderator_id):
    """Delete a moderator with confirmation and safety checks"""
    moderator = get_object_or_404(Moderator, id=moderator_id)
    
    # Check if moderator has assigned patients
    assigned_patients_count = Patient.objects.filter(moderator_assigned=moderator).count()
    
    if request.method == 'POST':
        if assigned_patients_count > 0:
            messages.error(request, 
                f'Cannot delete moderator "{moderator.user.get_full_name()}" because they have {assigned_patients_count} assigned patients. Please reassign these patients first.', extra_tags="admin")
            return redirect('moderator_list')
        
        try:
            moderator_name = moderator.user.get_full_name()
            user = moderator.user
            
            # Delete the moderator (this will also delete the user due to CASCADE)
            moderator.delete()
            user.delete()
            
            messages.success(request, f'Moderator "{moderator_name}" deleted successfully.', extra_tags="admin")
            return redirect('moderator_list')
            
        except Exception as e:
            messages.error(request, f'Error deleting moderator: {str(e)}', extra_tags="admin")
            return redirect('moderator_list')
    
    context = {
        'moderator': moderator,
        'assigned_patients_count': assigned_patients_count,
        'can_delete': assigned_patients_count == 0
    }
    
    return render(request, 'admin/moderator_confirm_delete.html', context)

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def doctor_list(request):
    """Display list of all doctors with search and filtering capabilities"""
    try:
        # Get search query from request
        search_query = request.GET.get('search', '').strip()
        
        # Start with all doctors
        doctors = Doctor.objects.select_related('user').all()
        
        # Apply search filter if provided
        if search_query:
            doctors = doctors.filter(
                models.Q(user__first_name__icontains=search_query) |
                models.Q(user__last_name__icontains=search_query) |
                models.Q(user__email__icontains=search_query) |
                models.Q(user__username__icontains=search_query) |
                models.Q(phone_number__icontains=search_query) |
                models.Q(specialization__icontains=search_query)
            )
        
        # Order by username for consistent display
        doctors = doctors.order_by('user__username')
        
        # Add escalated patient count for each doctor
        doctors_with_counts = []
        for doctor in doctors:
            escalated_patient_count = Patient.objects.filter(doctor_escalated=doctor).count()
            doctors_with_counts.append({
                'doctor': doctor,
                'escalated_patient_count': escalated_patient_count
            })
        
        context = {
            'doctors_with_counts': doctors_with_counts,
            'search_query': search_query,
            'total_count': len(doctors_with_counts)
        }
        
        return render(request, 'admin/doctor_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading doctor list: {str(e)}', extra_tags="admin")
        return redirect('admin_dashboard')

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def doctor_create(request):
    """Create a new doctor with form handling and validation"""
    if request.method == 'POST':
        from .forms import DoctorForm
        form = DoctorForm(request.POST)
        
        if form.is_valid():
            try:
                # Create the User object
                user = User.objects.create(
                    username=form.cleaned_data['username'],
                    email=form.cleaned_data['email'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name']
                )
                user.set_password(form.cleaned_data['password'])
                user.save()
                
                # Create the Doctor object
                doctor = Doctor.objects.create(
                    user=user,
                    phone_number=form.cleaned_data['phone_number'],
                    specialization=form.cleaned_data['specialization']
                )
                
                messages.success(request, f'Doctor "{user.get_full_name()}" created successfully.', extra_tags="admin")
                return redirect('doctor_list')
                
            except Exception as e:
                messages.error(request, f'Error creating doctor: {str(e)}', extra_tags="admin")
                # Clean up user if doctor creation failed
                if 'user' in locals():
                    user.delete()
        else:
            messages.error(request, 'Please correct the errors below.', extra_tags="admin")
    else:
        from .forms import DoctorForm
        form = DoctorForm()
    
    context = {
        'form': form,
        'title': 'Add New Doctor',
        'action': 'Create'
    }
    
    return render(request, 'admin/doctor_form.html', context)

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def doctor_detail(request, doctor_id):
    """Display detailed information about a specific doctor"""
    try:
        doctor = get_object_or_404(Doctor, id=doctor_id)
        
        # Get escalated patients for this doctor
        escalated_patients = Patient.objects.filter(doctor_escalated=doctor).select_related('user', 'moderator_assigned')
        
        # Format patient data for display
        formatted_patients = []
        for patient in escalated_patients:
            formatted_patients.append({
                'patient': patient,
                'age': patient.age,
                'moderator_name': patient.moderator_assigned.user.get_full_name() if patient.moderator_assigned else 'Not assigned'
            })
        
        context = {
            'doctor': doctor,
            'escalated_patients': formatted_patients,
            'escalated_count': len(formatted_patients)
        }
        
        return render(request, 'admin/doctor_detail.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading doctor details: {str(e)}', extra_tags="admin")
        return redirect('doctor_list')

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def doctor_edit(request, doctor_id):
    """Edit an existing doctor with form handling and validation"""
    doctor = get_object_or_404(Doctor, id=doctor_id)
    
    if request.method == 'POST':
        from .forms import DoctorForm
        form = DoctorForm(request.POST, instance=doctor)
        
        if form.is_valid():
            try:
                # Update the User object
                user = doctor.user
                user.username = form.cleaned_data['username']
                user.email = form.cleaned_data['email']
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                
                # Update password if provided
                if form.cleaned_data['password']:
                    user.set_password(form.cleaned_data['password'])
                
                user.save()
                
                # Update the Doctor object
                doctor.phone_number = form.cleaned_data['phone_number']
                doctor.specialization = form.cleaned_data['specialization']
                doctor.save()
                
                messages.success(request, f'Doctor "{user.get_full_name()}" updated successfully.', extra_tags="admin")
                return redirect('doctor_list')
                
            except Exception as e:
                messages.error(request, f'Error updating doctor: {str(e)}', extra_tags="admin")
        else:
            messages.error(request, 'Please correct the errors below.', extra_tags="admin")
    else:
        from .forms import DoctorForm
        form = DoctorForm(instance=doctor)
    
    context = {
        'form': form,
        'title': f'Edit Doctor: {doctor.user.get_full_name()}',
        'action': 'Update',
        'doctor': doctor
    }
    
    return render(request, 'admin/doctor_form.html', context)

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def doctor_delete(request, doctor_id):
    """Delete a doctor with confirmation and safety checks"""
    doctor = get_object_or_404(Doctor, id=doctor_id)
    
    # Check if doctor has escalated patients
    escalated_patients_count = Patient.objects.filter(doctor_escalated=doctor).count()
    
    if request.method == 'POST':
        if escalated_patients_count > 0:
            messages.error(request, 
                f'Cannot delete doctor "{doctor.user.get_full_name()}" because they have {escalated_patients_count} escalated patients. Please reassign these patients first.', extra_tags="admin")
            return redirect('doctor_list')
        
        try:
            doctor_name = doctor.user.get_full_name()
            user = doctor.user
            
            # Delete the doctor (this will also delete the user due to CASCADE)
            doctor.delete()
            user.delete()
            
            messages.success(request, f'Doctor "{doctor_name}" deleted successfully.', extra_tags="admin")
            return redirect('doctor_list')
            
        except Exception as e:
            messages.error(request, f'Error deleting doctor: {str(e)}', extra_tags="admin")
            return redirect('doctor_list')
    
    context = {
        'doctor': doctor,
        'escalated_patients_count': escalated_patients_count,
        'can_delete': escalated_patients_count == 0
    }
    
    return render(request, 'admin/doctor_confirm_delete.html', context)

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_patient_list(request):
    """Display list of all patients in the system for admin access"""
    try:
        # Get search query from request
        search_query = request.GET.get('search', '').strip()
        
        # Start with all patients
        patients = Patient.objects.select_related('user', 'moderator_assigned', 'doctor_escalated').all()
        
        # Apply search filter if provided
        if search_query:
            patients = patients.filter(
                models.Q(user__first_name__icontains=search_query) |
                models.Q(user__last_name__icontains=search_query) |
                models.Q(user__email__icontains=search_query) |
                models.Q(phone_number__icontains=search_query) |
                models.Q(insurance__icontains=search_query) |
                models.Q(monitoring_parameters__icontains=search_query)
            )
        
        # Order by creation date (newest first)
        patients = patients.order_by('-created_at')
        
        # Format patient data for display
        formatted_patients = []
        for patient in patients:
            # Format medications into a list if they exist
            medications = []
            if patient.medications:
                medications = [med.strip() for med in patient.medications.split('\n') if med.strip()]

            # Format pharmacy info into structured data if it exists
            pharmacy_info = {}
            if patient.pharmacy_info:
                try:
                    lines = patient.pharmacy_info.split('\n')
                    for line in lines:
                        if ':' in line:
                            key, value = line.split(':', 1)
                            pharmacy_info[key.strip()] = value.strip()
                except:
                    pharmacy_info = {'Details': patient.pharmacy_info}

            # Format allergies into a list if they exist
            allergies = []
            if patient.allergies:
                allergies = [allergy.strip() for allergy in patient.allergies.split(',') if allergy.strip()]

            # Format family history into structured sections if it exists
            family_history = []
            if patient.family_history:
                history_lines = patient.family_history.split('\n')
                for line in history_lines:
                    if line.strip():
                        family_history.append(line.strip())

            formatted_patient = {
                'patient': patient,
                'formatted_medications': medications,
                'formatted_pharmacy': pharmacy_info,
                'formatted_allergies': allergies,
                'formatted_family_history': family_history,
                'moderator_name': patient.moderator_assigned.user.get_full_name() if patient.moderator_assigned else 'Not assigned',
                'doctor_name': patient.doctor_escalated.user.get_full_name() if patient.doctor_escalated else 'Not escalated'
            }
            formatted_patients.append(formatted_patient)
        
        context = {
            'patient_obj': formatted_patients,
            'search_query': search_query,
            'total_count': len(formatted_patients),
            'is_admin_view': True
        }
        
        return render(request, 'admin/admin_patient_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading patient list: {str(e)}')
        return redirect('admin_dashboard')

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_patient_detail(request, patient_id):
    """Display detailed patient information for admin access"""
    try:
        patient = get_object_or_404(Patient, id=patient_id)
        
        # Get all reports for this patient
        reports = Reports.objects.filter(patient=patient).order_by('-created_at')
        
        # Get all doctors for potential escalation
        doctors = Doctor.objects.all()
        
        context = {
            'patient': patient,
            'reports': reports,
            'doctors': doctors,
            'is_admin_view': True
        }
        
        return render(request, 'index.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading patient details: {str(e)}')
        return redirect('admin_patient_list')

def express_interest(request):
    """Handle the express interest form for potential RPM leads"""
    # Get PMH choices for the form's select field
    pmh_choices = PastMedicalHistory.PMH_CHOICES
    
    if request.method == 'POST':
        # Process form submission
        form_data = {
            'first_name': request.POST.get('first_name'),
            'last_name': request.POST.get('last_name'),
            'email': request.POST.get('email'),
            'phone_number': request.POST.get('phone_number'),
            'date_of_birth': request.POST.get('date_of_birth'),
            'age': request.POST.get('age'),
            'sex': request.POST.get('sex'),
            'height': request.POST.get('height') or None,
            'weight': request.POST.get('weight') or None,
            'home_address': request.POST.get('home_address'),
            'emergency_contact_name': request.POST.get('emergency_contact_name'),
            'emergency_contact_phone': request.POST.get('emergency_contact_phone'),
            'emergency_contact_relationship': request.POST.get('emergency_contact_relationship'),
            'primary_care_physician': request.POST.get('primary_care_physician'),
            'primary_care_physician_phone': request.POST.get('primary_care_physician_phone'),
            'primary_care_physician_email': request.POST.get('primary_care_physician_email'),
            'insurance': request.POST.get('insurance'),
            'insurance_number': request.POST.get('insurance_number'),
            'device_serial_number': request.POST.get('device_serial_number'),
            'allergies': request.POST.get('allergies'),
            'medications': request.POST.get('medications'),
            'pharmacy_info': request.POST.get('pharmacy_info'),
            'family_history': request.POST.get('family_history'),
            'smoke': request.POST.get('smoke'),
            'drink': request.POST.get('drink'),
            'service_interest': request.POST.get('service_interest'),
            'additional_comments': request.POST.get('additional_comments'),
            'good_eyesight': 'good_eyesight' in request.POST,
            'can_follow_instructions': 'can_follow_instructions' in request.POST,
            'can_take_readings': 'can_take_readings' in request.POST,
            'medical_summary_file': request.FILES.get('medical_summary_file'),
        }
        
        # Get past medical history selections
        past_medical_history = request.POST.getlist('past_medical_history')
        
        try:
            # Create the Interest record
            interest = Interest.objects.create(**form_data)
            
            # Add past medical history entries
            for pmh in past_medical_history:
                InterestPastMedicalHistory.objects.create(
                    interest=interest,
                    pmh=pmh
                )
            
            messages.success(request, "Thank you for your interest! We'll be in touch soon.")
            return redirect('home')
            
        except Exception as e:
            messages.error(request, f"There was an error processing your form: {str(e)}")
            return render(request, 'express_interest.html', {
                'pmh_choices': pmh_choices,
                'form_data': form_data
            })
    
    # GET request - display the form
    return render(request, 'express_interest.html', {
        'pmh_choices': pmh_choices,
        'form_data': {}
    })

@api_view(["POST"])
@permission_classes([])
@authentication_classes([])
def create_patient(request):
    # check if the patient is already in the sso, and if the patient is in the sso, then check if the patient is already in the database and if the user in the database then return the patient object and if the user is not in the database then create new patient object, and if the user is not in sso then create the user in the sso and also in the database

    data = request.data
    email = data.get("email")
    password = data.get("password")
    first_name = data.get("first_name")
    last_name = data.get("last_name")
    entered_otp = data.get("entered_otp")
    phone_number = data.get("phone_number")
    date_of_birth = data.get('date_of_birth')
    height = data.get('height')
    weight = data.get('weight')
    
    print("Number")

    # check if the patient is already in the sso
    response = requests.post("https://auth.pinksurfing.com/api/signup/", data=data)
    if response.status_code == 409:
        return Response(
            {"error": response.text},
            status=status.HTTP_409_CONFLICT,
        )
    elif response.status_code != 201:
        return Response(
            {"error": response.text},
            status=status.HTTP_400_BAD_REQUEST,
        )
    else:
        # check if the patient is already in the database
        user = User.objects.filter(username = email).first()
        patient = Patient.objects.filter(user=user).first()
        if patient:
            return Response(
                {"message": "User already exists"}, status=status.HTTP_409_CONFLICT
            )
        else:
            # create new patient object
            user = User.objects.create(
                username=email,
                first_name=first_name,
                last_name=last_name,
                email=email,  # Use the same email as in the SSO for the local user
            )
            user.set_password(password)  # Properly hash the password
            user.save()
            patient = Patient.objects.create(
                user = user,
                date_of_birth=date_of_birth, 
                height=height,
                weight=weight,
            )
            return Response(
                {"message": "Patient successfully created"},
                status=status.HTTP_201_CREATED,
            )

@api_view(['POST'])            
def assign_patient(request, patient_id, moderator_id=""):
    patient = Patient.objects.get(id=patient)
    moderator = Moderator.objects.get(id=moderator_id)
    patient.moderator_assigned = moderator
    patient.save()
    return Response({"success": True}, status=status.HTTP_200_OK)
    
def moderator_actions_view(request, patient_id):
    # Get the moderator based on the email from the request (ensure authentication handles this)
    
    user = request.user  # Assuming user is authenticated and their email is available in request.user
    moderator = get_object_or_404(Moderator, user=user)
    patient = get_object_or_404(Patient, id=patient_id)

    context = {
        'patient': patient,
        'moderator': moderator,
    }
    action = request.GET.get('action')

    if action == 'access'.lower() and patient.moderator_assigned == moderator or request.user.is_superuser:
        # Access patient details
        reports = Reports.objects.filter(patient=patient)
        context['reports'] = reports  # Add patient details to the context
        context['doctors'] = Doctor.objects.all()  # Pass all doctors for escalation dropdown
        print(context['doctors'], "DEBUG: Doctors available for escalation")
        return render(request, 'index.html', context)

    elif action == 'update'.lower():
        if request.method == 'POST':
            # Handle form submission for updating the patient
            form = PatientForm(request.POST, instance=patient)
            if form.is_valid():
                form.save()
                context['patient_details'] = patient  # Updated patient
                context['success'] = 'Patient updated successfully'
            else:
                context['form_errors'] = form.errors
        else:
            # Prepopulate the form with patient data for rendering
            form = PatientForm(instance=patient)
            context['form'] = form

        return render(request, 'index.html', context)

    elif action == 'delete':
        # Delete patient
        patient.delete()
        context['success'] = 'Patient deleted successfully'
        return render(request, 'index.html', context)

    elif action == 'update-report':
        report_id = request.POST.get('report_id')
        report = get_object_or_404(Reports, id=report_id)

        if report.patient != patient:
            context['error'] = "This report does not belong to this patient."
            return render(request, 'index.html', context)

        if request.method == 'POST':
            # Handle report update
            form = ReportForm(request.POST, instance=report)
            if form.is_valid():
                form.save()
                context['success'] = "Report updated successfully"
                context['report'] = report
            else:
                context['form_errors'] = form.errors
        else:
            # Prepopulate the report form for updating
            form = ReportForm(instance=report)
            context['form'] = form

        return render(request, 'index.html', context)

    elif action == 'delete-report':
        report_id = request.POST.get('report_id')
        report = get_object_or_404(Reports, id=report_id)

        if report.patient != patient:
            context['error'] = "This report does not belong to this patient."
        else:
            report.delete()
            context['success'] = "Report deleted successfully"

        return render(request, 'index.html', context)

    else:
        # Invalid action
        context['error'] = "Invalid action."
        return render(request, 'index.html', context)

    
def all_moderators(request):
    moderators = Moderator.objects.all()
    serializer = ModeratorSerializer(moderators, many=True)
    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_patient(request, patient_id):
    pass

def moderator_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        # Authenticate the user
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Check if the user is a moderator
            if hasattr(user, 'moderator'):
                login(request, user)
                return redirect('view_all_assigned_patient')  # Redirect to moderator dashboard or home page
            else:
                messages.error(request, 'You are not authorized as a moderator.')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'moderator_login.html')

def moderator_logout(request):
    logout(request)  # Logs out the current user
    return redirect('/')  # Redirect to the login page after logout

@login_required
def view_assigned_patient(request):
    moderator = Moderator.objects.get(user=request.user)
    # Order patients by signup date (created_at)
    patient_obj = Patient.objects.filter(moderator_assigned=moderator).order_by('created_at')
    
    # Import models for call status
    from retell_calling.models import RetellCallSession, CallSummary
    
    formatted_patients = []
    serial_number = 1  # Start serial number from 1
    for patient in patient_obj:
        # Format medications into a list if they exist
        medications = []
        if patient.medications:
            medications = [med.strip() for med in patient.medications.split('\n') if med.strip()]

        # Format pharmacy info into structured data if it exists
        pharmacy_info = {}
        if patient.pharmacy_info:
            try:
                lines = patient.pharmacy_info.split('\n')
                for line in lines:
                    if ':' in line:
                        key, value = line.split(':', 1)
                        pharmacy_info[key.strip()] = value.strip()
            except:
                pharmacy_info = {'Details': patient.pharmacy_info}

        # Format allergies into a list if they exist
        allergies = []
        if patient.allergies:
            allergies = [allergy.strip() for allergy in patient.allergies.split(',') if allergy.strip()]

        # Format family history into structured sections if it exists
        family_history = []
        if patient.family_history:
            history_lines = patient.family_history.split('\n')
            for line in history_lines:
                if line.strip():
                    family_history.append(line.strip())

        # Get last documentation for this patient
        last_documentation = Documentation.objects.filter(patient=patient).order_by('-created_at').first()
        last_doc_text = None
        if last_documentation:
            last_doc_text = f"{last_documentation.title} - {timezone.localtime(last_documentation.created_at).strftime('%m/%d/%Y %I:%M %p')}"
        
        # Get last vital (report) for this patient
        last_vital = Reports.objects.filter(patient=patient).order_by('-created_at').first()
        last_vital_text = None
        if last_vital:
            vital_parts = []
            if last_vital.systolic_blood_pressure and last_vital.diastolic_blood_pressure:
                vital_parts.append(f"BP: {last_vital.systolic_blood_pressure}/{last_vital.diastolic_blood_pressure}")
            if last_vital.pulse:
                vital_parts.append(f"HR: {last_vital.pulse}")
            if last_vital.blood_glucose:
                vital_parts.append(f"BG: {last_vital.blood_glucose}")
            if last_vital.spo2:
                vital_parts.append(f"SpO2: {last_vital.spo2}%")
            
            if vital_parts:
                last_vital_text = f"{', '.join(vital_parts)} - {timezone.localtime(last_vital.created_at).strftime('%m/%d/%Y %I:%M %p')}"
            else:
                last_vital_text = f"Vital recorded - {timezone.localtime(last_vital.created_at).strftime('%m/%d/%Y %I:%M %p')}"

        # Get latest call status for this patient
        latest_call = RetellCallSession.objects.filter(patient=patient).order_by('-created_at').first()
        call_status = latest_call.call_status if latest_call else None
        
        # Get summary count for this patient
        summary_count = CallSummary.objects.filter(patient=patient).count()

        formatted_patient = {
            'serial_number': serial_number,
            'patient': patient,
            'formatted_medications': medications,
            'formatted_pharmacy': pharmacy_info,
            'formatted_allergies': allergies,
            'formatted_family_history': family_history,
            'last_documentation': last_doc_text,
            'last_vital': last_vital_text,
            'status': patient.status or 'green',
            'sticky_note': patient.sticky_note or '',
            'call_status': call_status,
            'summary_count': summary_count
        }
        formatted_patients.append(formatted_patient)
        serial_number += 1  # Increment serial number

    context = {
        'patient_obj': formatted_patients,
    }
    return render(request, 'view_assigned_patient.html', context)

@login_required
def write_document(request, report_id):
    report = Reports.objects.get(id=report_id)
    if request.method == 'POST':
        form = DocumentationForm(request.POST, request.FILES)
        if form.is_valid():
            # Pass the current user to the form's save method to set the 'written_by' field
            document = form.save(commit=False, user=request.user)
            document.report = report
            document.save()
            messages.success(request, 'Documentation saved successfully.')
            # Redirect to the view documentation page for this patient
            return redirect('view_documentation', patient_id=report.patient.id)
    else:
        form = DocumentationForm()
    
    # Get the patient related to the report to pass to the template
    patient = report.patient
    
    return render(request, 'reports/add_docs.html', {'form': form, 'report': report, 'patient': patient}) # Render add_docs.html

@login_required
def view_documentation(request, patient_id):
    # Check if the user is a moderator
    user = request.user
    is_moderator = Moderator.objects.filter(user=user).exists()
    patient = Patient.objects.filter(id=patient_id).first()
    is_patient = Patient.objects.filter(user=user, id=patient_id).exists()
    print(f"DEBUG: Patient = {patient}")
    
    # Allow access if user is a moderator or if the user is the patient
    if not is_moderator and not is_patient:
        print("DEBUG: User is not authorized to view documentation")
        return JsonResponse({"error": "You are not authorized to view this documentation"}, status=403)
      # Get documentation for the patient
    patient = get_object_or_404(Patient, id=patient_id)
    
    documentations = Documentation.objects.filter(
        patient=patient,
    ).order_by('-doc_report_date')
    documentation_list = []
    for doc in documentations:
        print("fldsaf",doc.doc_report_date)
        doc_data = {
            'id': doc.id,
            'title': doc.title,
            'history_of_present_illness': doc.history_of_present_illness,
            'chief_complaint': doc.chief_complaint,
            'subjective': doc.subjective,
            'objective': doc.objective,
            'assessment': doc.assessment,
            'plan': doc.plan,
            'written_by': doc.written_by,
            'created_at': timezone.localtime(doc.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            'updated_at': timezone.localtime(doc.updated_at).strftime("%Y-%m-%d %H:%M:%S"),
            'doc_report_date': doc.doc_report_date.strftime("%m/%d/%Y") if doc.doc_report_date else None,
        }
        
        # Only add file_url if the file exists
        if doc.file and hasattr(doc.file, 'url'):
            doc_data['file_url'] = doc.file.url
        else:
            doc_data['file_url'] = None
            
        documentation_list.append(doc_data)
        print(documentation_list)
    print('DEBUG: Documentation List:', documentation_list)

    return JsonResponse({'documentations': documentation_list})


@login_required
def register_patient(request):
    if request.method == 'POST':
        data = {
            'email': request.POST.get('email'),
            'password': request.POST.get('password'),
            'first_name': request.POST.get('first_name'), 
            'last_name': request.POST.get('last_name'),
            'phone_number': request.POST.get('phone_number'),
            'date_of_birth': request.POST.get('date_of_birth'),
            'height': request.POST.get('height'),
            'weight': request.POST.get('weight'),
            'insurance': request.POST.get('insurance'),
            'insurance_number': request.POST.get('insurance_number'),
            'sex': request.POST.get('sex'),
            'monitoring_parameters': request.POST.get('monitoring_parameters'),
            'device_serial_number': request.POST.get('device_serial_number'),
            'pharmacy_info': request.POST.get('pharmacy_info'),
            'allergies': request.POST.get('allergies'),
            'drink': request.POST.get('drink', 'NO'),
            'smoke': request.POST.get('smoke', 'NO'),
            'family_history': request.POST.get('family_history'),
            'medications': request.POST.get('medications'),
            'past_medical_history': request.POST.getlist('past_medical_history', []),
            'home_address': request.POST.get('home_address'),
            'emergency_contact_name': request.POST.get('emergency_contact_name'),
            'emergency_contact_phone': request.POST.get('emergency_contact_phone'),
            'emergency_contact_relationship': request.POST.get('emergency_contact_relationship'),
            'primary_care_physician': request.POST.get('primary_care_physician'),
            'primary_care_physician_phone': request.POST.get('primary_care_physician_phone')
        }

        # Check if patient already exists
        # if User.objects.filter(username=data['email']).exists():
        #     messages.error(request, 'Patient with this email already exists')
        #     return render(request, 'register_patient.html')

        # Create local user and patient atomically
        try:
            with transaction.atomic():
                user, created = User.objects.get_or_create(
                    username=data['email'],
                    defaults={
                        'email': data['email'],
                        'first_name': data['first_name'],
                        'last_name': data['last_name'],
                    }
                )

                if created:
                    user.set_password(data['password'])  # Properly hash the password
                    user.save()

                if Patient.objects.filter(user=user).exists():
                    messages.error(request, 'Patient with this email already exists')
                    return render(request, 'register_patient.html')
                
                patient = Patient.objects.create(
                    user=user,
                    date_of_birth=data['date_of_birth'],
                    height=data['height'] if data['height'] else 0.0,
                    weight=data['weight'] if data['weight'] else 0.0,
                    insurance=data['insurance'],
                    insurance_number=data['insurance_number'],
                    sex=data['sex'],
                    phone_number=data['phone_number'],
                    monitoring_parameters=data['monitoring_parameters'],
                    device_serial_number=data['device_serial_number'] if data['device_serial_number'] else None,
                    pharmacy_info=data['pharmacy_info'] if data['pharmacy_info'] else None,
                    allergies=data['allergies'] if data['allergies'] else None,
                    drink=data['drink'],
                    smoke=data['smoke'],
                    family_history=data['family_history'] if data['family_history'] else None,
                    medications=data['medications'] if data['medications'] else None,
                    home_address=data['home_address'] if data['home_address'] else None,
                    emergency_contact_name=data['emergency_contact_name'] if data['emergency_contact_name'] else None,
                    emergency_contact_phone=data['emergency_contact_phone'] if data['emergency_contact_phone'] else None,
                    emergency_contact_relationship=data['emergency_contact_relationship'] if data['emergency_contact_relationship'] else None,
                    primary_care_physician=data['primary_care_physician'] if data['primary_care_physician'] else None,
                    primary_care_physician_phone=data['primary_care_physician_phone'] if data['primary_care_physician_phone'] else None
                )

                # Create past medical history records
                for pmh in data['past_medical_history']:
                    PastMedicalHistory.objects.create(
                        patient=patient,
                        pmh=pmh
                    )

                # Assign the current moderator
                moderator = Moderator.objects.get(user=request.user)
                patient.moderator_assigned = moderator
                patient.save()

            messages.success(request, 'Patient registered successfully')
            return redirect('view_all_assigned_patient')

        except ValueError as e:
            messages.error(request, str(e))  # This will catch the BMI calculation error
            return render(request, 'register_patient.html')
        except Exception as e:
            messages.error(request, f'Error creating patient: {str(e)}')
            return render(request, 'register_patient.html')

    context = {
        'sex_choices': Patient.SEX_CHOICES,
        'monitoring_choices': Patient.MONITORING_CHOICES,
        'pmh_choices': PastMedicalHistory.PMH_CHOICES
    }
    return render(request, 'register_patient.html', context)

def registration_success(request):
    return render(request, 'registration_success.html')

@ratelimit(key="ip", rate="1/m", method='POST', block=True)
def send_email_otp(request):
    """Send OTP to email for verification"""
    if request.method == 'POST':
        email = request.POST.get('email')
        
        if not email:
            messages.error(request, 'Email is required')
            context = {
                'sex_choices': Patient.SEX_CHOICES,
                'monitoring_choices': Patient.MONITORING_CHOICES,
                'pmh_choices': PastMedicalHistory.PMH_CHOICES,
                'verified_email': None
            }
            return render(request, 'patient_self_register.html', context)
        
        # Check if user already exists
        if User.objects.filter(email=email).exists():
            messages.error(request, 'An account with this email already exists')
            context = {
                'sex_choices': Patient.SEX_CHOICES,
                'monitoring_choices': Patient.MONITORING_CHOICES,
                'pmh_choices': PastMedicalHistory.PMH_CHOICES,
                'verified_email': None
            }
            return render(request, 'patient_self_register.html', context)
        
        # Generate OTP
        otp_code = generate_otp()
        expires_at = timezone.now() + timedelta(minutes=10)
        
        # Delete any existing OTPs for this email
        EmailOTP.objects.filter(email=email).delete()
        
        # Create new OTP record
        otp_record = EmailOTP.objects.create(
            email=email,
            otp_code=otp_code,
            expires_at=expires_at
        )
        
        # Send OTP email
        if send_otp_email(email, otp_code):
            messages.success(request, f'Verification code sent to {email}')
            return render(request, 'verify_email_otp.html', {'email': email})
        else:
            messages.error(request, 'Failed to send verification email. Please try again.')
            context = {
                'sex_choices': Patient.SEX_CHOICES,
                'monitoring_choices': Patient.MONITORING_CHOICES,
                'pmh_choices': PastMedicalHistory.PMH_CHOICES,
                'verified_email': None
            }
            return render(request, 'patient_self_register.html', context)
    
    context = {
        'sex_choices': Patient.SEX_CHOICES,
        'monitoring_choices': Patient.MONITORING_CHOICES,
        'pmh_choices': PastMedicalHistory.PMH_CHOICES,
        'verified_email': None
    }
    return render(request, 'patient_self_register.html', context)

@ratelimit(key="ip", rate="1/m", method='POST', block=False)
def verify_email_otp(request):
    """Verify OTP and proceed to registration form"""
    print(f"DEBUG: verify_email_otp called with method: {request.method}")
    
    if request.method == 'POST':
        email = request.POST.get('email')
        otp_code = request.POST.get('otp_code')
        
        print(f"DEBUG: Received OTP verification - Email: {email}, OTP: {otp_code}")
        
        if not email or not otp_code:
            messages.error(request, 'Email and OTP code are required')
            return render(request, 'verify_email_otp.html', {'email': email})
        
        # Find the OTP record
        try:
            print(f"DEBUG: Looking for OTP record with email: {email}, code: {otp_code}")
            
            # Test if EmailOTP table exists
            try:
                EmailOTP.objects.count()
                print(f"DEBUG: EmailOTP table exists")
            except Exception as table_error:
                print(f"DEBUG: EmailOTP table error: {str(table_error)}")
                messages.error(request, 'Database error. Please contact support.')
                return render(request, 'verify_email_otp.html', {'email': email})
            
            otp_record = EmailOTP.objects.filter(
                email=email,
                otp_code=otp_code,
                is_verified=False
            ).latest('created_at')
            
            print(f"DEBUG: Found OTP record: {otp_record}")
            
            # Check if OTP is expired
            if otp_record.is_expired():
                print(f"DEBUG: OTP expired")
                messages.error(request, 'OTP code has expired. Please request a new one.')
                return render(request, 'verify_email_otp.html', {'email': email})
            
            # Mark OTP as verified
            otp_record.is_verified = True
            otp_record.save()
            
            # Store email in session for registration form
            request.session['verified_email'] = email
            
            print(f"DEBUG: OTP verified successfully for {email}, redirecting to patient_register")
            messages.success(request, 'Email verified successfully! Please complete your registration.')
            return redirect('patient_register')
            
        except EmailOTP.DoesNotExist:
            print(f"DEBUG: OTP record not found")
            messages.error(request, 'Invalid OTP code. Please check and try again.')
            return render(request, 'verify_email_otp.html', {'email': email})
        except Exception as e:
            print(f"DEBUG: Error in OTP verification: {str(e)}")
            messages.error(request, f'Error verifying OTP: {str(e)}')
            return render(request, 'verify_email_otp.html', {'email': email})
    
    # Handle GET request - redirect to registration if no email provided
    return redirect('patient_register')

# @ratelimit(key="ip", rate="1/m", method='POST', block=True)
def patient_self_registration(request):
    verified_email = request.session.get('verified_email')
    
    if request.method == 'POST':
        # Check if email is verified in session for POST requests
        if not verified_email:
            messages.error(request, 'Please verify your email first')
            return redirect('patient_register')
        data = {
            'email': verified_email,  # Use verified email from session
            'password': request.POST.get('password'),
            'first_name': request.POST.get('first_name'), 
            'last_name': request.POST.get('last_name'),
            'phone_number': request.POST.get('phone_number'),
            'date_of_birth': request.POST.get('date_of_birth'),
            'height': request.POST.get('height'),
            'weight': request.POST.get('weight'),
            'insurance': request.POST.get('insurance'),
            'insurance_number': request.POST.get('insurance_number'),
            'sex': request.POST.get('sex'),
            'monitoring_parameters': request.POST.get('monitoring_parameters'),
            'device_serial_number': request.POST.get('device_serial_number'),
            'pharmacy_info': request.POST.get('pharmacy_info'),
            'allergies': request.POST.get('allergies'),
            'drink': request.POST.get('drink', 'NO'),
            'smoke': request.POST.get('smoke', 'NO'),
            'family_history': request.POST.get('family_history'),
            'medications': request.POST.get('medications'),
            'past_medical_history': request.POST.getlist('past_medical_history', []),
            'home_address': request.POST.get('home_address'),
            'emergency_contact_name': request.POST.get('emergency_contact_name'),
            'emergency_contact_phone': request.POST.get('emergency_contact_phone'),
            'emergency_contact_relationship': request.POST.get('emergency_contact_relationship'),
            'primary_care_physician': request.POST.get('primary_care_physician'),
            'primary_care_physician_phone': request.POST.get('primary_care_physician_phone')
        }
        try:
            with transaction.atomic():
                user, created = User.objects.get_or_create(
                    username=data['email'],
                    defaults={
                        'email': data['email'],
                        'first_name': data['first_name'],
                        'last_name': data['last_name'],
                    }
                )
                if created:
                    user.set_password(data['password'])  # Properly hash the password
                    user.save()

                if Patient.objects.filter(user=user).exists():
                    messages.error(request, 'Patient with this email already exists')
                    return render(request, 'register_patient.html')

                patient = Patient.objects.create(
                    user=user,
                    date_of_birth=data['date_of_birth'],
                    height=data['height'] if data['height'] else 0.0,
                    weight=data['weight'] if data['weight'] else 0.0,
                    insurance=data['insurance'],
                    insurance_number=data['insurance_number'],
                    sex=data['sex'],
                    phone_number=data['phone_number'],
                    monitoring_parameters=data['monitoring_parameters'],
                    device_serial_number=data['device_serial_number'] if data['device_serial_number'] else None,
                    pharmacy_info=data['pharmacy_info'] if data['pharmacy_info'] else None,
                    allergies=data['allergies'] if data['allergies'] else None,
                    drink=data['drink'],
                    smoke=data['smoke'],
                    family_history=data['family_history'] if data['family_history'] else None,
                    medications=data['medications'] if data['medications'] else None,
                    home_address=data['home_address'] if data['home_address'] else None,
                    emergency_contact_name=data['emergency_contact_name'] if data['emergency_contact_name'] else None,
                    emergency_contact_phone=data['emergency_contact_phone'] if data['emergency_contact_phone'] else None,
                    emergency_contact_relationship=data['emergency_contact_relationship'] if data['emergency_contact_relationship'] else None,
                    primary_care_physician=data['primary_care_physician'] if data['primary_care_physician'] else None,
                    primary_care_physician_phone=data['primary_care_physician_phone'] if data['primary_care_physician_phone'] else None
                )

                # Create past medical history records
                for pmh in data['past_medical_history']:
                    PastMedicalHistory.objects.create(
                        patient=patient,
                        pmh=pmh
                    )

            messages.success(request, 'Patient registered successfully')
            # Clear verified email from session
            if 'verified_email' in request.session:
                del request.session['verified_email']
            return redirect('registration_success')

        except ValueError as e:
            messages.error(request, str(e))  # This will catch the BMI calculation error
            context = {
                'sex_choices': Patient.SEX_CHOICES,
                'monitoring_choices': Patient.MONITORING_CHOICES,
                'pmh_choices': PastMedicalHistory.PMH_CHOICES,
                'verified_email': verified_email
            }
            return render(request, 'patient_self_register.html', context)
        except Exception as e:
            messages.error(request, f'Error creating patient: {str(e)}')
            context = {
                'sex_choices': Patient.SEX_CHOICES,
                'monitoring_choices': Patient.MONITORING_CHOICES,
                'pmh_choices': PastMedicalHistory.PMH_CHOICES,
                'verified_email': verified_email
            }
            return render(request, 'patient_self_register.html', context)

    context = {
        'sex_choices': Patient.SEX_CHOICES,
        'monitoring_choices': Patient.MONITORING_CHOICES,
        'pmh_choices': PastMedicalHistory.PMH_CHOICES,
        'verified_email': verified_email
    }
    return render(request, 'patient_self_register.html', context)


@login_required
def view_patient(request, patient_id):
    patient = Patient.objects.get(id=patient_id)
    
    # Process allergies
    allergies_list = []
    if patient.allergies:
        allergies_list = [allergy.strip() for allergy in patient.allergies.split(',')]
    
    # Process medications
    medications_list = []
    if patient.medications:
        medications_list = [med.strip() for med in patient.medications.split('\n')]
    
    # Process family history
    family_history_list = []
    if patient.family_history:
        family_history_list = [history.strip() for history in patient.family_history.split('\n')]
    
    # Process pharmacy info
    pharmacy_info_list = []
    if patient.pharmacy_info:
        pharmacy_info_list = [info.strip() for info in patient.pharmacy_info.split('\n')]
    
    context = {
        'patient': patient,
        'allergies_list': allergies_list,
        'medications_list': medications_list,
        'family_history_list': family_history_list,
        'pharmacy_info_list': pharmacy_info_list,
    }
    return render(request, 'index.html', context)

@never_cache
@login_required
def admin_logout(request):
    """Custom admin logout view that handles both GET and POST requests"""
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('home')

def patient_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        print(f"Login attempt - Username: {username}")
        print(f"Login attempt - Password: {password}")
        
        user = authenticate(request, username=username, password=password)
        print(f"User: {user}")  # Debug print
        
        if user is not None:
            # Check if user is a moderator
            if Moderator.objects.filter(user=user).exists():
                print("User is a moderator, redirecting to moderator login")
                return JsonResponse({
                    'success': False,
                    'error': 'Please use the moderator login page'
                })
            
            # Check if user is a patient
            try:
                patient = Patient.objects.get(user=user)
                print(f"Found patient: {patient}")  # Debug print
                login(request, user)
                return JsonResponse({
                    'success': True,
                    'redirect_url': '/patient_home/'
                })
            except Patient.DoesNotExist:
                print("User exists but no associated patient found")  # Debug print
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid credentials'
                })
        else:
            print("Authentication failed - user is None")  # Debug print
            return JsonResponse({
                'success': False,
                'error': 'Invalid credentials'
            })
    
    return render(request, 'reports/patient_login.html')

@login_required
def patient_home(request):
    try:
        patient = Patient.objects.get(user=request.user)
        reports = Reports.objects.filter(patient=patient).order_by('-created_at')
        context = {
            'patient': patient,
            'reports': reports
        }
        return render(request, 'reports/patient_home.html', context)
    except Patient.DoesNotExist:
        logout(request)
        return redirect('patient_login')

def patient_logout(request):
    logout(request)
    return redirect('home')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        print(f"User: {user}")  # Debug print
        if user is not None:
            login(request, user)
            return redirect('admin:index')
        else:
            # Handle invalid login
            return render(request, 'admin/login.html', {'error': 'Invalid username or password'})
    return render(request, 'admin/login.html')

@csrf_exempt
def track_interest(request):
    """Enhanced API endpoint for tracking partial lead data with session-based tracking"""
    if request.method == "POST":
        import logging
        import re
        from datetime import datetime
        from django.utils import timezone
        from django.db import transaction
        
        logger = logging.getLogger(__name__)
        
        try:
            # Ensure the user has a session for anonymous tracking
            if not request.session.session_key:
                request.session.save()
            session_key = request.session.session_key
            
            # Log the tracking attempt for monitoring
            logger.info(f"Lead tracking attempt for session: {session_key}")
            
            # Parse and validate JSON data
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError as e:
                logger.error(f"Invalid JSON in track_interest: {str(e)}, session: {session_key}")
                return JsonResponse({"success": False, "error": "Invalid JSON data"}, status=400)
            
            # Enhanced data validation and sanitization
            validated_data = {}
            validation_errors = []
            
            # Define comprehensive field validation rules
            field_rules = {
                "first_name": {"type": str, "max_length": 100, "sanitize": True},
                "last_name": {"type": str, "max_length": 100, "sanitize": True},
                "email": {"type": str, "max_length": 254, "validate": "email", "sanitize": True},
                "phone_number": {"type": str, "max_length": 20, "validate": "phone", "sanitize": True},
                "date_of_birth": {"type": str, "validate": "date"},
                "age": {"type": int, "min": 0, "max": 150},
                "allergies": {"type": str, "max_length": 1000, "sanitize": True},
                "service_interest": {"type": str, "max_length": 50, "sanitize": True},
                "insurance": {"type": str, "max_length": 255, "sanitize": True},
                "additional_comments": {"type": str, "max_length": 2000, "sanitize": True},
                "good_eyesight": {"type": bool},
                "can_follow_instructions": {"type": bool},
                "can_take_readings": {"type": bool},
                "sex": {"type": str, "max_length": 10, "sanitize": True},
                "height": {"type": str, "max_length": 10, "sanitize": True},
                "weight": {"type": str, "max_length": 10, "sanitize": True},
                "home_address": {"type": str, "max_length": 500, "sanitize": True},
                "emergency_contact_name": {"type": str, "max_length": 255, "sanitize": True},
                "emergency_contact_phone": {"type": str, "max_length": 20, "validate": "phone", "sanitize": True},
                "emergency_contact_relationship": {"type": str, "max_length": 100, "sanitize": True},
                "primary_care_physician": {"type": str, "max_length": 255, "sanitize": True},
                "primary_care_physician_phone": {"type": str, "max_length": 20, "validate": "phone", "sanitize": True},
                "primary_care_physician_email": {"type": str, "max_length": 254, "validate": "email", "sanitize": True},
                "insurance_number": {"type": str, "max_length": 255, "sanitize": True},
                "device_serial_number": {"type": str, "max_length": 255, "sanitize": True},
                "medications": {"type": str, "max_length": 2000, "sanitize": True},
                "pharmacy_info": {"type": str, "max_length": 1000, "sanitize": True},
                "family_history": {"type": str, "max_length": 2000, "sanitize": True},
                "smoke": {"type": str, "max_length": 10, "sanitize": True},
                "drink": {"type": str, "max_length": 10, "sanitize": True},
            }
            
            # Validate and sanitize each field
            for field, value in data.items():
                if field not in field_rules:
                    logger.warning(f"Unknown field '{field}' in request, session: {session_key}")
                    continue  # Skip unknown fields
                    
                rules = field_rules[field]
                
                # Skip empty values but log them for monitoring
                if value is None or value == "":
                    logger.debug(f"Empty value for field '{field}', session: {session_key}")
                    continue
                
                try:
                    # Type conversion and validation
                    if rules["type"] == str:
                        value = str(value)
                        
                        # Enhanced sanitization
                        if rules.get("sanitize"):
                            # Remove potentially harmful characters and normalize whitespace
                            value = re.sub(r'[<>"\']', '', value)  # Remove HTML/script chars
                            value = re.sub(r'\s+', ' ', value)     # Normalize whitespace
                            value = value.strip()
                        
                        if len(value) == 0:
                            continue
                            
                        # Length validation with truncation
                        if "max_length" in rules and len(value) > rules["max_length"]:
                            logger.warning(f"Field '{field}' truncated from {len(value)} to {rules['max_length']} chars, session: {session_key}")
                            value = value[:rules["max_length"]]
                        
                        # Enhanced email validation
                        if rules.get("validate") == "email":
                            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                            if not re.match(email_pattern, value):
                                logger.warning(f"Invalid email format: {value}, session: {session_key}")
                                validation_errors.append(f"Invalid email format: {field}")
                                continue
                        
                        # Phone number validation
                        if rules.get("validate") == "phone":
                            # Remove non-digit characters for validation
                            phone_digits = re.sub(r'[^\d]', '', value)
                            if len(phone_digits) < 10 or len(phone_digits) > 15:
                                logger.warning(f"Invalid phone format: {value}, session: {session_key}")
                                validation_errors.append(f"Invalid phone format: {field}")
                                continue
                    
                    elif rules["type"] == int:
                        try:
                            value = int(float(value))  # Handle string numbers
                        except (ValueError, TypeError):
                            logger.warning(f"Invalid integer value for '{field}': {value}, session: {session_key}")
                            validation_errors.append(f"Invalid number format: {field}")
                            continue
                            
                        if "min" in rules and value < rules["min"]:
                            logger.warning(f"Value too small for '{field}': {value}, session: {session_key}")
                            validation_errors.append(f"Value too small: {field}")
                            continue
                        if "max" in rules and value > rules["max"]:
                            logger.warning(f"Value too large for '{field}': {value}, session: {session_key}")
                            validation_errors.append(f"Value too large: {field}")
                            continue
                    
                    elif rules["type"] == bool:
                        if isinstance(value, bool):
                            pass  # Already boolean
                        elif isinstance(value, str):
                            value = value.lower() in ['true', '1', 'yes', 'on']
                        else:
                            value = bool(value)
                    
                    # Enhanced date validation
                    if rules.get("validate") == "date":
                        try:
                            # Try multiple date formats
                            date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']
                            parsed_date = None
                            
                            for date_format in date_formats:
                                try:
                                    parsed_date = datetime.strptime(value, date_format).date()
                                    break
                                except ValueError:
                                    continue
                            
                            if not parsed_date:
                                raise ValueError("No valid date format found")
                            
                            # Validate date is reasonable (not in future, not too old)
                            today = timezone.now().date()
                            if parsed_date > today:
                                logger.warning(f"Future date provided for '{field}': {value}, session: {session_key}")
                                validation_errors.append(f"Future date not allowed: {field}")
                                continue
                            
                            # Convert to standard format
                            value = parsed_date.strftime('%Y-%m-%d')
                            
                        except ValueError as e:
                            logger.warning(f"Invalid date format for '{field}': {value}, session: {session_key}")
                            validation_errors.append(f"Invalid date format: {field}")
                            continue
                    
                    validated_data[field] = value
                    
                except (ValueError, TypeError) as e:
                    logger.warning(f"Validation error for field '{field}': {str(e)}, session: {session_key}")
                    validation_errors.append(f"Validation error: {field}")
                    continue
            
            # Log validation results
            if validation_errors:
                logger.info(f"Validation errors for session {session_key}: {', '.join(validation_errors)}")
            
            # Use database transaction for data consistency
            with transaction.atomic():
                # Enhanced duplicate handling - look for existing leads by session or email
                lead = None
                
                # First try to find by session key
                if session_key:
                    lead = InterestLead.objects.filter(session_key=session_key).first()
                
                # If no lead found by session and we have an email, check for existing email
                if not lead and validated_data.get('email'):
                    existing_email_lead = InterestLead.objects.filter(
                        email=validated_data['email']
                    ).first()
                    
                    if existing_email_lead:
                        # Merge session tracking with existing email lead
                        existing_email_lead.session_key = session_key
                        existing_email_lead.save()
                        lead = existing_email_lead
                        logger.info(f"Merged session {session_key} with existing email lead {lead.id}")
                
                # Create new lead if none found
                if not lead:
                    lead = InterestLead.objects.create(session_key=session_key)
                    logger.info(f"Created new lead {lead.id} with session_key: {session_key}")
                else:
                    logger.info(f"Updating existing lead {lead.id} with session_key: {session_key}")
                
                # Enhanced data merging - preserve existing data unless explicitly overwritten
                updated_fields = []
                for field, value in validated_data.items():
                    current_value = getattr(lead, field, None)
                    
                    # Special handling for different field types
                    if field == "date_of_birth":
                        if value and value != "":
                            try:
                                # Convert string to date object for comparison
                                new_date = datetime.strptime(value, '%Y-%m-%d').date()
                                if current_value != new_date:
                                    setattr(lead, field, new_date)
                                    updated_fields.append(field)
                            except ValueError:
                                logger.error(f"Failed to parse date {value} for lead {lead.id}")
                    
                    elif field == "age":
                        if value is not None and current_value != value:
                            setattr(lead, field, value)
                            updated_fields.append(field)
                    
                    else:
                        # For other fields, only update if new value is different and not empty
                        if value is not None and value != "" and current_value != value:
                            setattr(lead, field, value)
                            updated_fields.append(field)
                
                # Save the lead with updated timestamp and validation
                try:
                    lead.save()
                    
                    # Enhanced logging for monitoring
                    if updated_fields:
                        logger.info(f"Updated lead {lead.id} fields: {', '.join(updated_fields)}, completion: {lead.completion_percentage}%")
                    else:
                        logger.info(f"No changes for lead {lead.id}, completion: {lead.completion_percentage}%")
                        
                except Exception as validation_error:
                    logger.error(f"Validation error saving lead {lead.id}: {str(validation_error)}")
                    return JsonResponse({
                        'success': False,
                        'error': f'Validation error: {str(validation_error)}',
                        'lead_id': str(lead.id) if lead.id else None
                    }, status=400)
                
                # Calculate completion metrics for response
                completion_percentage = lead.completion_percentage
                is_complete = lead.is_complete
                
                # Log completion milestones
                if completion_percentage >= 50 and completion_percentage < 75:
                    logger.info(f"Lead {lead.id} reached 50%+ completion")
                elif completion_percentage >= 75 and completion_percentage < 100:
                    logger.info(f"Lead {lead.id} reached 75%+ completion")
                elif is_complete:
                    logger.info(f"Lead {lead.id} is now complete and ready for conversion")
                
                # Return enhanced success response
                response_data = {
                    "success": True,
                    "lead_id": str(lead.id),
                    "session_key": session_key,
                    "updated_fields": updated_fields,
                    "completion_percentage": completion_percentage,
                    "is_complete": is_complete,
                    "validation_errors": validation_errors if validation_errors else None,
                    "timestamp": timezone.now().isoformat()
                }
                
                return JsonResponse(response_data)
            
        except Exception as e:
            # Comprehensive error logging with context
            import traceback
            error_details = traceback.format_exc()
            error_context = {
                "session_key": session_key if 'session_key' in locals() else 'unknown',
                "data_keys": list(data.keys()) if 'data' in locals() else 'unknown',
                "error": str(e),
                "traceback": error_details
            }
            
            logger.error(f"Critical error in track_interest: {error_context}")
            
            # Return user-friendly error message
            return JsonResponse({
                "success": False, 
                "error": "An error occurred while saving your information. Please try again.",
                "timestamp": timezone.now().isoformat()
            }, status=500)
    
    # Handle non-POST requests
    logger.warning(f"Invalid method {request.method} for track_interest endpoint")
    return JsonResponse({"error": "Only POST method is allowed"}, status=405)

def terms_and_conditions_view(request):
    """Renders the terms and conditions page."""
    return render(request, 'terms_and_conditions.html')

@login_required
def edit_documentation(request, doc_id):
    documentation = Documentation.objects.get(id=doc_id)
    report = documentation.report if hasattr(documentation, 'report') else None
    if request.method == 'POST':
        form = DocumentationForm(request.POST, request.FILES, instance=documentation)
        if form.is_valid():
            document = form.save(commit=False, user=request.user)
            if report:
                document.report = report
            document.save()
            messages.success(request, 'Documentation updated successfully.')
            return redirect('view_documentation', patient_id=documentation.patient.id)
    else:
        form = DocumentationForm(instance=documentation, initial={
            'full_documentation': documentation.history_of_present_illness
        })
    patient = documentation.patient
    return render(request, 'reports/add_docs.html', {'form': form, 'reports': report, 'patient': patient, 'edit_mode': True, 'doc_id': doc_id})

def doctor_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and Doctor.objects.filter(user=user).exists():
            login(request, user)
            return redirect('view_escalated_patient')
        else:
            messages.error(request, 'Invalid credentials or not a doctor.')
    return render(request, 'doctor_login.html')

@login_required
def view_escalated_patient(request):
    # Only allow doctors
    try:
        doctor = Doctor.objects.get(user=request.user)
    except Doctor.DoesNotExist:
        logout(request)
        return redirect('doctor_login')
    # List all escalated patients assigned to this doctor
    patients = Patient.objects.filter(doctor_escalated=doctor, is_escalated=True)
    formatted_patients = []
    for patient in patients:
        medications = [med.strip() for med in patient.medications.split('\n')] if patient.medications else []
        allergies = [allergy.strip() for allergy in patient.allergies.split(',')] if patient.allergies else []
        family_history = [fh.strip() for fh in patient.family_history.split('\n')] if patient.family_history else []
        pharmacy_info = [info.strip() for info in patient.pharmacy_info.split('\n')] if patient.pharmacy_info else []
        formatted_patients.append({
            'patient': patient,
            'formatted_medications': medications,
            'formatted_allergies': allergies,
            'formatted_family_history': family_history,
            'formatted_pharmacy': pharmacy_info,
        })
    context = {
        'patients': formatted_patients,
    }
    print(context, "DEBUG: Escalated patients context")
    return render(request, 'view_escalated_patient.html', context)

@login_required
def doctor_patient_detail(request, patient_id):
    try:
        doctor = Doctor.objects.get(user=request.user)
    except Doctor.DoesNotExist:
        logout(request)
        return redirect('doctor_login')
    patient = get_object_or_404(Patient, id=patient_id, doctor_escalated=doctor, is_escalated=True)

    # Format medications into a list if they exist
    medications = []
    if patient.medications:
        medications = [med.strip() for med in patient.medications.split('\n') if med.strip()]

    # Format pharmacy info into structured data if it exists
    pharmacy_info = {}
    if patient.pharmacy_info:
        try:
            lines = patient.pharmacy_info.split('\n')
            for line in lines:
                if ':' in line:
                    key, value = line.split(':', 1)
                    pharmacy_info[key.strip()] = value.strip()
        except:
            pharmacy_info = {'Details': patient.pharmacy_info}

    # Format allergies into a list if they exist
    allergies = []
    if patient.allergies:
        allergies = [allergy.strip() for allergy in patient.allergies.split(',') if allergy.strip()]

    # Format family history into structured sections if it exists
    family_history = []
    if patient.family_history:
        history_lines = patient.family_history.split('\n')
        for line in history_lines:
            if line.strip():
                family_history.append(line.strip())

    context = {
        'patient': patient,
        'formatted_medications': medications,
        'formatted_pharmacy': pharmacy_info,
        'formatted_allergies': allergies,
        'formatted_family_history': family_history
    }
    return render(request, 'index.html', context)

def doctor_logout(request):
    logout(request)
    return redirect('doctor_login')



# Escalation logic
@login_required
@require_POST
def escalate_patient(request, patient_id):
    moderator = get_object_or_404(Moderator, user=request.user)
    patient = get_object_or_404(Patient, id=patient_id, moderator_assigned=moderator)
    doctor_id = request.POST.get('doctor_id')
    doctor = get_object_or_404(Doctor, id=doctor_id)
    patient.doctor_escalated = doctor
    patient.is_escalated = True
    patient.save()

    # SendGrid email notification to doctor
    sg = sendgrid.SendGridAPIClient(api_key=settings.SENDGRID_API_KEY)
    doctor_email = doctor.user.email
    patient_user = patient.user
    message = Mail(
        from_email='marketing@pinksurfing.com',
        to_emails=doctor_email,
        subject='A Patient Has Been Escalated to You',
        html_content=f"""
        <h3>Patient Escalation Notification</h3>
        <p>Dear Dr. {doctor.user.get_full_name() or doctor.user.username},</p>
        <p>The following patient has been escalated to you:</p>
        <ul>
            <li><strong>Name:</strong> {patient_user.first_name} {patient_user.last_name}</li>
            <li><strong>Email:</strong> {patient_user.email}</li>
            <li><strong>Date of Birth:</strong> {patient.date_of_birth}</li>
            <li><strong>Sex:</strong> {patient.sex}</li>
            <li><strong>Phone Number:</strong> {patient.phone_number}</li>
            <li><strong>Insurance:</strong> {patient.insurance}</li>
            <li><strong>Allergies:</strong> {patient.allergies or "N/A"}</li>
            <li><strong>Medications:</strong> {patient.medications or "N/A"}</li>
            <li><strong>Family History:</strong> {patient.family_history or "N/A"}</li>
        </ul>
        <p>Please log in to the portal to view more details.</p>
        """
    )
    try:
        sg.send(message)
    except Exception as e:
        print("SendGrid error:", e)

    messages.success(request, f"Patient escalated to Dr. {doctor.user.get_full_name()}")
    return redirect('moderator_actions', patient_id=patient.id)

# Admin-verified user creation views
def admin_create_user(request):
    """Display form for creating moderator or doctor accounts with admin verification"""
    return render(request, 'admin_create_user.html')

def test_staff_ui(request):
    """Test page for staff creation workflow"""
    with open('/app/test_staff_ui.html', 'r') as f:
        content = f.read()
    return HttpResponse(content, content_type='text/html')

@csrf_exempt
def verify_admin_password(request):
    """Verify admin password before allowing user creation"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admin_password = data.get('admin_password')
            
            # Get the first superuser for authentication
            admin_user = User.objects.filter(is_superuser=True).first()
            
            if not admin_user:
                return JsonResponse({
                    'success': False,
                    'error': 'No admin user found in the system.'
                })
            
            # Check if the provided password is correct
            if admin_user.check_password(admin_password):
                return JsonResponse({
                    'success': True,
                    'message': 'Admin password verified successfully.'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid admin password.'
                })
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data.'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method.'
    })

def create_staff_user(request):
    """Create moderator or doctor account after admin verification"""
    print(f"DEBUG: create_staff_user called with method: {request.method}")
    print(f"DEBUG: Content type: {request.content_type}")
    
    if request.method == 'POST':
        try:
            print("DEBUG: Processing POST request for staff user creation")
            
            # Verify admin password first
            admin_password = request.POST.get('admin_password')
            print(f"DEBUG: Admin password provided: {'Yes' if admin_password else 'No'}")
            
            admin_user = User.objects.filter(is_superuser=True).first()
            
            if not admin_user or not admin_user.check_password(admin_password):
                print("DEBUG: Admin password verification failed")
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid admin password.'
                })
            
            print("DEBUG: Admin password verified successfully")
            
            # Get form data
            user_type = request.POST.get('user_type')  # 'moderator' or 'doctor'
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            phone_number = request.POST.get('phone_number')
            
            print(f"DEBUG: Form data - user_type: {user_type}, username: {username}, email: {email}")
            
            # Additional fields for doctor
            specialization = request.POST.get('specialization') if user_type == 'doctor' else None
            
            # Validation
            if not all([user_type, username, email, password, first_name, last_name]):
                return JsonResponse({
                    'success': False,
                    'error': 'All required fields must be filled.'
                })
            
            if user_type not in ['moderator', 'doctor']:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid user type.'
                })
            
            # Check if username already exists
            if User.objects.filter(username=username).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Username already exists.'
                })
            
            # Check if email already exists
            if User.objects.filter(email=email).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Email already exists.'
                })
            
            # Create user
            print(f"DEBUG: Creating {user_type} user")
            if user_type == 'moderator':
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_staff=True  # Moderators are staff users
                )
                print(f"DEBUG: User created with ID: {user.id}")
                
                # Create moderator profile
                moderator = Moderator.objects.create(
                    user=user,
                    phone_number=phone_number or ''
                )
                print(f"DEBUG: Moderator profile created with ID: {moderator.id}")
                
                success_message = f'Moderator account created successfully for {first_name} {last_name}.'
                
            elif user_type == 'doctor':
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                print(f"DEBUG: Doctor user created with ID: {user.id}")
                
                # Create doctor profile
                doctor = Doctor.objects.create(
                    user=user,
                    specialization=specialization or '',
                    phone_number=phone_number or ''
                )
                print(f"DEBUG: Doctor profile created with ID: {doctor.id}")
                
                success_message = f'Doctor account created successfully for Dr. {first_name} {last_name}.'
            
            print(f"DEBUG: Returning success response: {success_message}")
            return JsonResponse({
                'success': True,
                'message': success_message
            })
            
        except Exception as e:
            print(f"DEBUG: Exception occurred: {str(e)}")
            print(f"DEBUG: Exception type: {type(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method.'
    })


# Moderator Shortcut API Views
@login_required
def get_shortcuts(request):
    """Get all shortcuts for the current moderator"""
    try:
        moderator = get_object_or_404(Moderator, user=request.user)
        from .models import ModeratorShortcut
        shortcuts = ModeratorShortcut.objects.filter(moderator=moderator)
        shortcuts_data = [
            {
                'id': shortcut.id,
                'shortcut_key': shortcut.shortcut_key,
                'description': shortcut.description,
                'content': shortcut.content
            }
            for shortcut in shortcuts
        ]
        return JsonResponse({'shortcuts': shortcuts_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def create_shortcut(request):
    """Create a new shortcut for the current moderator"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        moderator = get_object_or_404(Moderator, user=request.user)
        from .models import ModeratorShortcut
        
        data = json.loads(request.body)
        shortcut_key = data.get('shortcut_key', '').strip()
        description = data.get('description', '').strip()
        content = data.get('content', '').strip()
        
        if not all([shortcut_key, description, content]):
            return JsonResponse({'error': 'All fields are required'}, status=400)
        
        # Check if shortcut already exists for this moderator
        if ModeratorShortcut.objects.filter(moderator=moderator, shortcut_key=shortcut_key).exists():
            return JsonResponse({'error': 'Shortcut key already exists'}, status=400)
        
        shortcut = ModeratorShortcut.objects.create(
            moderator=moderator,
            shortcut_key=shortcut_key,
            description=description,
            content=content
        )
        
        return JsonResponse({
            'success': True,
            'shortcut': {
                'id': shortcut.id,
                'shortcut_key': shortcut.shortcut_key,
                'description': shortcut.description,
                'content': shortcut.content
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def update_shortcut(request, shortcut_id):
    """Update an existing shortcut"""
    if request.method != 'PUT':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        moderator = get_object_or_404(Moderator, user=request.user)
        from .models import ModeratorShortcut
        shortcut = get_object_or_404(ModeratorShortcut, id=shortcut_id, moderator=moderator)
        
        data = json.loads(request.body)
        shortcut_key = data.get('shortcut_key', '').strip()
        description = data.get('description', '').strip()
        content = data.get('content', '').strip()
        
        if not all([shortcut_key, description, content]):
            return JsonResponse({'error': 'All fields are required'}, status=400)
        
        # Check if new shortcut key conflicts with existing ones (excluding current)
        if (shortcut_key != shortcut.shortcut_key and 
            ModeratorShortcut.objects.filter(moderator=moderator, shortcut_key=shortcut_key).exists()):
            return JsonResponse({'error': 'Shortcut key already exists'}, status=400)
        
        shortcut.shortcut_key = shortcut_key
        shortcut.description = description
        shortcut.content = content
        shortcut.save()
        
        return JsonResponse({
            'success': True,
            'shortcut': {
                'id': shortcut.id,
                'shortcut_key': shortcut.shortcut_key,
                'description': shortcut.description,
                'content': shortcut.content
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def delete_shortcut(request, shortcut_id):
    """Delete a shortcut"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        moderator = get_object_or_404(Moderator, user=request.user)
        from .models import ModeratorShortcut
        shortcut = get_object_or_404(ModeratorShortcut, id=shortcut_id, moderator=moderator)
        shortcut.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def search_shortcuts(request):
    """Search shortcuts by key for autocomplete"""
    try:
        moderator = get_object_or_404(Moderator, user=request.user)
        from .models import ModeratorShortcut
        query = request.GET.get('q', '').strip()
        
        if not query:
            return JsonResponse({'shortcuts': []})
        
        shortcuts = ModeratorShortcut.objects.filter(
            moderator=moderator,
            shortcut_key__icontains=query
        )[:10]  # Limit to 10 results
        
        shortcuts_data = [
            {
                'shortcut_key': shortcut.shortcut_key,
                'description': shortcut.description,
                'content': shortcut.content
            }
            for shortcut in shortcuts
        ]
        return JsonResponse({'shortcuts': shortcuts_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def handle_excel_import(request):
    """Handle Excel file import for leads"""
    try:
        import openpyxl
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        import tempfile
        import os
        
        print("handle_excel_import function called")
        print("Request FILES:", request.FILES)
        excel_file = request.FILES['excel_file']
        print(f"Excel file received: {excel_file.name}, size: {excel_file.size}")
        
        # Validate file type - accept Excel and CSV files
        if not excel_file.name.lower().endswith(('.xlsx', '.xls', '.csv')):
            messages.error(request, 'Please upload a valid Excel or CSV file (.xlsx, .xls, or .csv)')
            return redirect('leads_list')
        
        # Determine file type and save accordingly
        file_name_lower = excel_file.name.lower()
        if file_name_lower.endswith('.csv'):
            file_extension = '.csv'
        elif file_name_lower.endswith('.xlsx'):
            file_extension = '.xlsx'
        else:
            file_extension = '.xls'
            
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        print(f"Temporary file saved as: {tmp_file_path}")
        
        try:
            workbook = None
            worksheet = None
            
            # Handle CSV files
            if file_extension == '.csv':
                print("Processing CSV file...")
                import csv
                
                # Read CSV file
                with open(tmp_file_path, 'r', encoding='utf-8') as csvfile:
                    csv_reader = csv.reader(csvfile)
                    rows = list(csv_reader)
                
                if not rows:
                    raise Exception("CSV file is empty")
                
                print(f"CSV file has {len(rows)} rows")
                print(f"Headers: {rows[0]}")
                
                # Convert CSV to openpyxl workbook for consistent processing
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                
                # Write CSV data to worksheet
                for row_idx, row in enumerate(rows, 1):
                    for col_idx, value in enumerate(row, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                print("Successfully converted CSV to Excel format")
            
            # Handle Excel files
            else:
                # First try: Use pandas if available (handles both .xlsx and .xls)
                try:
                    import pandas as pd
                    print("Trying to read with pandas...")
                    df = pd.read_excel(tmp_file_path)
                    print(f"Successfully read Excel file with pandas. Shape: {df.shape}")
                    print(f"Columns: {list(df.columns)}")
                    
                    # Convert to openpyxl workbook for processing
                    converted_path = tmp_file_path.replace(file_extension, '_converted.xlsx')
                    df.to_excel(converted_path, index=False, engine='openpyxl')
                    workbook = openpyxl.load_workbook(converted_path)
                    worksheet = workbook.active
                    
                    # Clean up converted file
                    os.unlink(converted_path)
                    print("Successfully processed with pandas")
                    
                except ImportError:
                    print("Pandas not available, trying openpyxl directly...")
                except Exception as pandas_error:
                    print(f"Pandas failed to read file: {pandas_error}")
                
                # Second try: Direct openpyxl (works for .xlsx files)
                if workbook is None:
                    try:
                        print("Trying openpyxl directly...")
                        workbook = openpyxl.load_workbook(tmp_file_path)
                        worksheet = workbook.active
                        print("Successfully processed with openpyxl")
                    except Exception as openpyxl_error:
                        print(f"Openpyxl failed: {openpyxl_error}")
                
                if workbook is None:
                    raise Exception("Could not read Excel file with any available method")
            
            # Get headers from first row
            headers = [cell.value for cell in worksheet[1]]
            
            # Find all available columns (case-insensitive)
            column_mapping = {}
            
            for i, header in enumerate(headers, 1):
                if header and isinstance(header, str):
                    header_lower = header.lower().strip()
                    
                    # Map Excel columns to model fields (case-insensitive matching)
                    if 'first' in header_lower and 'name' in header_lower:
                        column_mapping['first_name'] = i
                    elif 'last' in header_lower and 'name' in header_lower:
                        column_mapping['last_name'] = i
                    elif header_lower == 'phone_number' or (header_lower == 'phone' and 'phone 2' not in headers[i:i+2]):
                        column_mapping['phone_number'] = i
                    elif header_lower == 'phone_number_2':
                        column_mapping['phone_number_2'] = i
                    elif header_lower == 'street 1':
                        column_mapping['street_address'] = i
                    elif header_lower == 'city':
                        column_mapping['city'] = i
                    elif header_lower == 'zip code':
                        column_mapping['zip_code'] = i
                    elif header_lower == 'mrn#' or 'mrn' in header_lower:
                        column_mapping['mrn_number'] = i
                    elif header_lower == 'date' or 'date' in header_lower:
                        column_mapping['date_of_birth'] = i
                    elif header_lower == 'sex' or header_lower == 's':
                        column_mapping['sex'] = i
                    elif header_lower == 'marital status' or header_lower == 'm':
                        column_mapping['marital_status'] = i
                    elif header_lower == 'primary insured id':
                        column_mapping['primary_insured_id'] = i
                    elif header_lower == 'primary insurance':
                        column_mapping['insurance'] = i
            
            # Debug: Print column mapping
            print(f"Column mapping found: {column_mapping}")
            print(f"Available headers: {headers}")
            
            # Check for required columns
            if not column_mapping.get('first_name') or not column_mapping.get('last_name') or not column_mapping.get('phone_number'):
                messages.error(request, 'Excel file must contain columns for first_name, last_name, and phone_number')
                return redirect('leads_list')
            
            # Process rows
            created_count = 0
            skipped_count = 0
            errors = []
            
            for row_num in range(2, worksheet.max_row + 1):  # Skip header row
                try:
                    # Extract all available data from the row
                    row_data = {}
                    
                    for field_name, col_index in column_mapping.items():
                        cell_value = worksheet.cell(row=row_num, column=col_index).value
                        if cell_value is not None:
                            row_data[field_name] = str(cell_value).strip()
                        else:
                            row_data[field_name] = ''
                    
                    # Get required fields
                    first_name = row_data.get('first_name', '')
                    last_name = row_data.get('last_name', '')
                    phone_number = row_data.get('phone_number', '')
                    
                    # Skip empty rows
                    if not first_name and not last_name and not phone_number:
                        continue
                    
                    print(f"Processing row {row_num}: {first_name} {last_name}, phone: {phone_number}")
                    
                    # Skip if no phone number (required field)
                    if not phone_number:
                        skipped_count += 1
                        errors.append(f'Row {row_num}: Missing phone number')
                        continue
                    
                    # Skip if no first_name AND no last_name (at least one required)
                    if not first_name and not last_name:
                        skipped_count += 1
                        errors.append(f'Row {row_num}: Either first name or last name must be provided')
                        continue
                    
                    # Check if lead already exists (by phone number)
                    if InterestLead.objects.filter(phone_number=phone_number).exists():
                        skipped_count += 1
                        errors.append(f'Row {row_num}: Lead with phone number {phone_number} already exists')
                        continue
                    
                    # Parse date of birth if available
                    date_of_birth = None
                    if row_data.get('date_of_birth'):
                        try:
                            from datetime import datetime
                            # Handle different date formats (MM/DD, MM/DD/YYYY, etc.)
                            date_str = row_data['date_of_birth']
                            if '/' in date_str:
                                parts = date_str.split('/')
                                if len(parts) == 2:  # MM/DD format
                                    # Assume current year for MM/DD format
                                    current_year = datetime.now().year
                                    date_of_birth = datetime.strptime(f"{date_str}/{current_year}", "%m/%d/%Y").date()
                                elif len(parts) == 3:  # MM/DD/YYYY format
                                    date_of_birth = datetime.strptime(date_str, "%m/%d/%Y").date()
                        except Exception as date_error:
                            print(f"Date parsing error for row {row_num}: {date_error}")
                    
                    # Create new lead with all available data
                    try:
                        lead_data = {
                            'first_name': first_name,
                            'last_name': last_name,
                            'phone_number': phone_number,
                        }
                        
                        # Add optional fields if they exist
                        if row_data.get('phone_number_2'):
                            lead_data['phone_number_2'] = row_data['phone_number_2']
                        if row_data.get('street_address'):
                            lead_data['street_address'] = row_data['street_address']
                        if row_data.get('city'):
                            lead_data['city'] = row_data['city']
                        if row_data.get('zip_code'):
                            lead_data['zip_code'] = row_data['zip_code']
                        if row_data.get('mrn_number'):
                            lead_data['mrn_number'] = row_data['mrn_number']
                        if date_of_birth:
                            lead_data['date_of_birth'] = date_of_birth
                        if row_data.get('sex'):
                            lead_data['sex'] = row_data['sex'].upper()[:1]  # Take first character and uppercase
                        if row_data.get('marital_status'):
                            lead_data['marital_status'] = row_data['marital_status']
                        if row_data.get('primary_insured_id'):
                            lead_data['primary_insured_id'] = row_data['primary_insured_id']
                        if row_data.get('insurance'):
                            lead_data['insurance'] = row_data['insurance']
                        
                        lead = InterestLead.objects.create(**lead_data)
                        created_count += 1
                        print(f"Created lead: {lead}")
                        
                    except Exception as validation_error:
                        skipped_count += 1
                        errors.append(f'Row {row_num}: Validation error - {str(validation_error)}')
                        continue
                    
                except Exception as e:
                    skipped_count += 1
                    errors.append(f'Row {row_num}: {str(e)}')
            
            # Clean up temporary file
            os.unlink(tmp_file_path)
            
            # Show results
            if created_count > 0:
                messages.success(request, f'Successfully imported {created_count} leads')
            
            if skipped_count > 0:
                messages.warning(request, f'Skipped {skipped_count} rows due to errors or duplicates')
            
            if errors and len(errors) <= 10:  # Show first 10 errors
                for error in errors[:10]:
                    messages.warning(request, error)
            elif len(errors) > 10:
                messages.warning(request, f'First 10 errors: {", ".join(errors[:10])}')
            
        except Exception as e:
            # Clean up temporary file
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
            messages.error(request, f'Error processing Excel file: {str(e)}')
            
    except ImportError:
        messages.error(request, 'Excel processing library not installed. Please install openpyxl.')
    except Exception as e:
        messages.error(request, f'Error importing leads: {str(e)}')
    
    return redirect('leads_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def delete_all_leads(request):
    """Delete all leads from the system"""
    import logging
    logger = logging.getLogger(__name__)
    
    if request.method == 'POST':
        try:
            # Get count before deletion for logging
            total_leads = InterestLead.objects.count()
            
            if total_leads == 0:
                messages.info(request, 'No leads to delete.')
                return redirect('leads_list')
            
            # Delete all leads (this will cascade to related objects)
            deleted_count = InterestLead.objects.all().delete()[0]
            
            messages.success(request, f'Successfully deleted {deleted_count} leads and all associated data.')
            logger.info(f'Admin {request.user.username} deleted all {deleted_count} leads')
            
        except Exception as e:
            messages.error(request, f'Error deleting leads: {str(e)}')
            logger.error(f'Error deleting all leads: {str(e)}')
    
    return redirect('leads_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def leads_list(request):
    """Display list of all leads with search and making patient capabilities"""
    try:
        from django.core.paginator import Paginator
        from datetime import datetime
        
        print(f"leads_list view called with method: {request.method}")
        
        # Handle Excel import if POST request
        if request.method == 'POST' and 'excel_file' in request.FILES:
            print("Excel import request detected")
            return handle_excel_import(request)
        elif request.method == 'POST':
            print("POST request but no excel_file in FILES")
            print("Available FILES:", request.FILES.keys())
            print("Available POST data:", request.POST.keys())
        
        # Get search and filter parameters
        search_query = request.GET.get('search', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        conversion_status = request.GET.get('conversion_status', '').strip()
        
        # Start with all leads
        leads = InterestLead.objects.all()
        
        # Apply search filter if provided
        if search_query:
            leads = leads.filter(
                models.Q(first_name__icontains=search_query) |
                models.Q(last_name__icontains=search_query) |
                models.Q(email__icontains=search_query) |
                models.Q(phone_number__icontains=search_query)
            )
        
        # Apply date range filter if provided
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                leads = leads.filter(created_at__date__gte=date_from_obj)
            except ValueError:
                messages.warning(request, 'Invalid "from" date format. Please use YYYY-MM-DD.')
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                leads = leads.filter(created_at__date__lte=date_to_obj)
            except ValueError:
                messages.warning(request, 'Invalid "to" date format. Please use YYYY-MM-DD.')
        
        # Apply conversion status filter if provided
        if conversion_status == 'converted':
            leads = leads.filter(is_converted=True)
        elif conversion_status == 'not_converted':
            leads = leads.filter(is_converted=False)
        
        # Order by creation date (newest first)
        leads = leads.order_by('-created_at')
        
        # Add computed fields for each lead
        leads_with_data = []
        for lead in leads:
            leads_with_data.append({
                'lead': lead,
                'completion_percentage': lead.completion_percentage,
                'is_complete': lead.is_complete,
                'converted_patient_name': lead.converted_patient.user.get_full_name() if lead.converted_patient else None,
                'converted_by_name': lead.converted_by.get_full_name() if lead.converted_by else None
            })
        
        # Implement pagination
        paginator = Paginator(leads_with_data, 25)  # Show 25 leads per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Calculate statistics
        total_leads = InterestLead.objects.count()
        converted_leads = InterestLead.objects.filter(is_converted=True).count()
        conversion_rate = round((converted_leads / total_leads * 100), 1) if total_leads > 0 else 0
        
        context = {
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': page_obj.has_other_pages(),
            'search_query': search_query,
            'date_from': date_from,
            'date_to': date_to,
            'conversion_status': conversion_status,
            'total_leads': total_leads,
            'converted_leads': converted_leads,
            'conversion_rate': conversion_rate,
            'filtered_count': len(leads_with_data)
        }
        
        return render(request, 'admin/leads_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading leads list: {str(e)}')
        return redirect('admin_dashboard')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def lead_detail(request, lead_id):
    """Display detailed information about a specific lead"""
    try:
        lead = get_object_or_404(InterestLead, id=lead_id)
        
        # Calculate completion status
        completion_percentage = lead.completion_percentage
        is_complete = lead.is_complete
        
        # Identify missing required fields for patient conversion
        missing_fields = []
        if not lead.first_name:
            missing_fields.append('First Name')
        if not lead.last_name:
            missing_fields.append('Last Name')
        if not lead.email:
            missing_fields.append('Email')
        if not lead.phone_number:
            missing_fields.append('Phone Number')
        if not lead.date_of_birth:
            missing_fields.append('Date of Birth')
        if not lead.insurance:
            missing_fields.append('Insurance')
        
        # Get conversion history if applicable
        conversion_history = None
        if lead.is_converted:
            conversion_history = {
                'converted_at': lead.converted_at,
                'converted_by': lead.converted_by,
                'converted_patient': lead.converted_patient
            }
        
        # Get all available fields and their values for display
        lead_fields = [
            {'label': 'First Name', 'value': lead.first_name, 'required': True},
            {'label': 'Last Name', 'value': lead.last_name, 'required': True},
            {'label': 'Email', 'value': lead.email, 'required': True},
            {'label': 'Phone Number', 'value': lead.phone_number, 'required': True},
            {'label': 'Date of Birth', 'value': lead.date_of_birth, 'required': True},
            {'label': 'Age', 'value': lead.age, 'required': False},
            {'label': 'Insurance', 'value': lead.insurance, 'required': True},
            {'label': 'Service Interest', 'value': lead.service_interest, 'required': False},
            {'label': 'Allergies', 'value': lead.allergies, 'required': False},
            {'label': 'Additional Comments', 'value': lead.additional_comments, 'required': False},
        ]
        
        context = {
            'lead': lead,
            'lead_fields': lead_fields,
            'completion_percentage': completion_percentage,
            'is_complete': is_complete,
            'missing_fields': missing_fields,
            'conversion_history': conversion_history,
            'can_convert': not lead.is_converted and is_complete
        }
        
        return render(request, 'admin/lead_detail.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading lead details: {str(e)}')
        return redirect('leads_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def convert_lead_to_patient(request, lead_id):
    """Convert a lead to a patient with pre-populated data"""
    try:
        from django.db import transaction
        from django.utils import timezone
        
        lead = get_object_or_404(InterestLead, id=lead_id)
        
        # Check if lead is already converted
        if lead.is_converted:
            messages.warning(request, 'This lead has already been converted to a patient.')
            return redirect('lead_detail', lead_id=lead_id)
        
        if request.method == 'POST':
            # Use transaction to ensure data consistency
            try:
                with transaction.atomic():
                    # Get form data with fallbacks to lead data
                    username = request.POST.get('username') or lead.email
                    email = request.POST.get('email') or lead.email
                    first_name = request.POST.get('first_name') or lead.first_name
                    last_name = request.POST.get('last_name') or lead.last_name
                    phone_number = request.POST.get('phone_number') or lead.phone_number
                    date_of_birth = request.POST.get('date_of_birth') or lead.date_of_birth
                    insurance = request.POST.get('insurance') or lead.insurance
                    allergies = request.POST.get('allergies') or lead.allergies
                    
                    # Additional patient-specific fields
                    height = request.POST.get('height')
                    weight = request.POST.get('weight')
                    sex = request.POST.get('sex')
                    monitoring_parameters = request.POST.get('monitoring_parameters')
                    device_serial_number = request.POST.get('device_serial_number')
                    pharmacy_info = request.POST.get('pharmacy_info')
                    smoke = request.POST.get('smoke', 'NO')
                    drink = request.POST.get('drink', 'NO')
                    family_history = request.POST.get('family_history')
                    medications = request.POST.get('medications')
                    
                    # Validate required fields
                    if not all([username, email, first_name, last_name, phone_number, date_of_birth, insurance]):
                        messages.error(request, 'Please fill in all required fields.')
                        return redirect('convert_lead_to_patient', lead_id=lead_id)
                    
                    # Check if user with this email already exists
                    if User.objects.filter(email=email).exists():
                        messages.error(request, f'A user with email {email} already exists.')
                        return redirect('convert_lead_to_patient', lead_id=lead_id)
                    
                    # Create User object
                    user = User.objects.create(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name
                    )
                    
                    # Set a temporary password (admin should inform patient to reset)
                    user.set_password('TempPassword123!')
                    user.save()
                    
                    # Create Patient object
                    patient_data = {
                        'user': user,
                        'date_of_birth': date_of_birth,
                        'phone_number': phone_number,
                        'insurance': insurance,
                        'sex': sex,
                        'monitoring_parameters': monitoring_parameters,
                        'pharmacy_info': pharmacy_info,
                        'allergies': allergies,
                        'smoke': smoke,
                        'drink': drink,
                        'family_history': family_history,
                        'medications': medications,
                    }
                    
                    # Add optional numeric fields if provided
                    if height:
                        try:
                            patient_data['height'] = float(height)
                        except ValueError:
                            pass
                    
                    if weight:
                        try:
                            patient_data['weight'] = float(weight)
                        except ValueError:
                            pass
                    
                    if device_serial_number:
                        try:
                            patient_data['device_serial_number'] = int(device_serial_number)
                        except ValueError:
                            pass
                    
                    patient = Patient.objects.create(**patient_data)
                    
                    # Update lead conversion status
                    lead.is_converted = True
                    lead.converted_patient = patient
                    lead.converted_at = timezone.now()
                    lead.converted_by = request.user
                    lead.save()
                    
                    messages.success(request, f'Lead successfully converted to patient: {patient.user.get_full_name()}')
                    return redirect('lead_detail', lead_id=lead_id)
                    
            except Exception as e:
                messages.error(request, f'Error converting lead to patient: {str(e)}')
                return redirect('convert_lead_to_patient', lead_id=lead_id)
        
        # GET request - display the conversion form
        # Pre-populate form with lead data
        form_data = {
            'username': lead.email,
            'email': lead.email,
            'first_name': lead.first_name,
            'last_name': lead.last_name,
            'phone_number': lead.phone_number,
            'date_of_birth': lead.date_of_birth,
            'insurance': lead.insurance,
            'allergies': lead.allergies,
        }
        
        # Get choices for dropdowns
        sex_choices = Patient.SEX_CHOICES
        monitoring_choices = Patient.MONITORING_CHOICES
        
        context = {
            'lead': lead,
            'form_data': form_data,
            'sex_choices': sex_choices,
            'monitoring_choices': monitoring_choices,
        }
        
        return render(request, 'admin/convert_lead_to_patient.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading conversion form: {str(e)}')
        return redirect('leads_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def delete_lead(request, lead_id):
    """Delete a lead with confirmation"""
    if request.method == 'POST':
        try:
            lead = get_object_or_404(InterestLead, id=lead_id)
            
            # Store lead info for success message
            lead_name = f"{lead.first_name or ''} {lead.last_name or ''}".strip()
            if not lead_name:
                lead_name = f"Lead {lead.id}"
            
            
            # Delete the lead
            lead.delete()
            
            messages.success(request, f'Lead "{lead_name}" has been successfully deleted.')
            return redirect('leads_list')
            
        except Exception as e:
            messages.error(request, f'Error deleting lead: {str(e)}')
            return redirect('lead_detail', lead_id=lead_id)
    
    # If not POST request, redirect to lead detail
    return redirect('lead_detail', lead_id=lead_id)


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def leads_call_summaries_list(request):
    """
    Display all leads with their call summary counts for admin access.
    """
    try:
        # Get all leads with their call summary counts
        leads_with_calls = InterestLead.objects.annotate(
            call_count=models.Count('lead_call_summaries'),
            latest_call=models.Max('lead_call_summaries__generated_at')
        ).filter(call_count__gt=0).order_by('-latest_call')
        
        # Get leads without calls
        leads_without_calls = InterestLead.objects.annotate(
            call_count=models.Count('lead_call_summaries')
        ).filter(call_count=0).order_by('first_name', 'last_name')
        
        # Calculate statistics for admin dashboard
        total_lead_calls = LeadCallSession.objects.count()
        leads_with_calls_count = leads_with_calls.count()
        
        context = {
            'leads_with_calls': leads_with_calls,
            'leads_without_calls': leads_without_calls,
            'total_summaries': LeadCallSummary.objects.count(),
            'total_leads_with_calls': leads_with_calls_count,
            'lead_calls_count': total_lead_calls,
            'leads_with_calls_count': leads_with_calls_count,
        }
        
        return render(request, 'retell_calling/leads_call_summaries_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading leads call summaries: {str(e)}')
        return redirect('admin_dashboard')


# ==================== Video Management Views ====================

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def video_list(request):
    """Display list of all videos for admin management"""
    try:
        from .models import Video
        
        # Get search query
        search_query = request.GET.get('search', '').strip()
        status_filter = request.GET.get('status', 'all')
        
        # Base queryset
        videos = Video.objects.all()
        
        # Apply search filter
        if search_query:
            videos = videos.filter(
                models.Q(title__icontains=search_query) |
                models.Q(description__icontains=search_query)
            )
        
        # Apply status filter
        if status_filter == 'active':
            videos = videos.filter(is_active=True)
        elif status_filter == 'inactive':
            videos = videos.filter(is_active=False)
        
        # Order by custom order field
        videos = videos.order_by('order', '-created_at')
        
        context = {
            'videos': videos,
            'total_videos': Video.objects.count(),
            'active_videos': Video.objects.filter(is_active=True).count(),
            'inactive_videos': Video.objects.filter(is_active=False).count(),
            'search_query': search_query,
            'status_filter': status_filter,
        }
        
        return render(request, 'videos/video_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading videos: {str(e)}')
        return redirect('admin_dashboard')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def video_create(request):
    """Create a new video"""
    try:
        from .models import Video
        
        if request.method == 'POST':
            title = request.POST.get('title', '').strip()
            youtube_url = request.POST.get('youtube_url', '').strip()
            description = request.POST.get('description', '').strip()
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            
            # Validation
            if not title:
                messages.error(request, 'Video title is required.')
                return redirect('video_create')
            
            if not youtube_url:
                messages.error(request, 'YouTube URL is required.')
                return redirect('video_create')
            
            # Validate YouTube URL format
            import re
            youtube_pattern = r'(youtube\.com|youtu\.be)'
            if not re.search(youtube_pattern, youtube_url):
                messages.error(request, 'Please enter a valid YouTube URL.')
                return redirect('video_create')
            
            # Verify it can extract a video ID
            video_id_patterns = [
                r'(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/)([a-zA-Z0-9_-]{11})',
                r'youtube\.com\/watch\?.*v=([a-zA-Z0-9_-]{11})',
                r'youtube\.com\/shorts\/([a-zA-Z0-9_-]{11})',
            ]
            
            video_id_found = False
            for pattern in video_id_patterns:
                if re.search(pattern, youtube_url):
                    video_id_found = True
                    break
            
            if not video_id_found:
                messages.error(request, 'Could not extract video ID from the YouTube URL. Please check the URL format.')
                return redirect('video_create')
            
            # Create video
            video = Video.objects.create(
                title=title,
                youtube_url=youtube_url,
                description=description,
                order=int(order) if order else 0,
                is_active=is_active,
                created_by=request.user
            )
            
            messages.success(request, f'Video "{title}" created successfully!')
            return redirect('video_list')
        
        # GET request - show form
        context = {
            'max_order': Video.objects.count(),
        }
        return render(request, 'videos/video_form.html', context)
        
    except Exception as e:
        messages.error(request, f'Error creating video: {str(e)}')
        return redirect('video_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def video_edit(request, video_id):
    """Edit an existing video"""
    try:
        from .models import Video
        
        video = get_object_or_404(Video, id=video_id)
        
        if request.method == 'POST':
            title = request.POST.get('title', '').strip()
            youtube_url = request.POST.get('youtube_url', '').strip()
            description = request.POST.get('description', '').strip()
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            
            # Validation
            if not title:
                messages.error(request, 'Video title is required.')
                return redirect('video_edit', video_id=video_id)
            
            if not youtube_url:
                messages.error(request, 'YouTube URL is required.')
                return redirect('video_edit', video_id=video_id)
            
            # Validate YouTube URL format
            import re
            youtube_pattern = r'(youtube\.com|youtu\.be)'
            if not re.search(youtube_pattern, youtube_url):
                messages.error(request, 'Please enter a valid YouTube URL.')
                return redirect('video_edit', video_id=video_id)
            
            # Verify it can extract a video ID
            video_id_patterns = [
                r'(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/)([a-zA-Z0-9_-]{11})',
                r'youtube\.com\/watch\?.*v=([a-zA-Z0-9_-]{11})',
                r'youtube\.com\/shorts\/([a-zA-Z0-9_-]{11})',
            ]
            
            video_id_found = False
            for pattern in video_id_patterns:
                if re.search(pattern, youtube_url):
                    video_id_found = True
                    break
            
            if not video_id_found:
                messages.error(request, 'Could not extract video ID from the YouTube URL. Please check the URL format.')
                return redirect('video_edit', video_id=video_id)
            
            # Update video
            video.title = title
            video.youtube_url = youtube_url
            video.description = description
            video.order = int(order) if order else 0
            video.is_active = is_active
            video.save()
            
            messages.success(request, f'Video "{title}" updated successfully!')
            return redirect('video_list')
        
        # GET request - show form
        context = {
            'video': video,
            'is_edit': True,
            'max_order': Video.objects.count(),
        }
        return render(request, 'videos/video_form.html', context)
        
    except Exception as e:
        messages.error(request, f'Error editing video: {str(e)}')
        return redirect('video_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def video_delete(request, video_id):
    """Delete a video"""
    try:
        from .models import Video
        
        video = get_object_or_404(Video, id=video_id)
        
        if request.method == 'POST':
            title = video.title
            video.delete()
            messages.success(request, f'Video "{title}" deleted successfully!')
            return redirect('video_list')
        
        # GET request - show confirmation
        context = {
            'video': video,
        }
        return render(request, 'videos/video_confirm_delete.html', context)
        
    except Exception as e:
        messages.error(request, f'Error deleting video: {str(e)}')
        return redirect('video_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
@csrf_exempt
def video_toggle_active(request, video_id):
    """Toggle video active status via AJAX"""
    try:
        from .models import Video
        
        if request.method == 'POST':
            video = get_object_or_404(Video, id=video_id)
            video.is_active = not video.is_active
            video.save()
            
            return JsonResponse({
                'success': True,
                'is_active': video.is_active,
                'message': f'Video is now {"active" if video.is_active else "inactive"}'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
@csrf_exempt
def video_reorder(request):
    """Update video order via AJAX"""
    try:
        from .models import Video
        
        if request.method == 'POST':
            data = json.loads(request.body)
            video_orders = data.get('video_orders', [])
            
            # Update each video's order
            for item in video_orders:
                video_id = item.get('id')
                new_order = item.get('order')
                
                video = Video.objects.get(id=video_id)
                video.order = new_order
                video.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Video order updated successfully'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# Public API for landing page
@api_view(['GET'])
@permission_classes([])
@authentication_classes([])
def get_active_videos(request):
    """Public API endpoint to get active videos for landing page"""
    try:
        from .models import Video
        import logging
        logger = logging.getLogger(__name__)
        
        videos = Video.objects.filter(is_active=True).order_by('order', '-created_at')
        
        video_list = []
        for video in videos:
            embed_url = video.get_embed_url()
            logger.info(f"Video: {video.title}, YouTube URL: {video.youtube_url}, Embed URL: {embed_url}")
            
            video_list.append({
                'id': str(video.id),
                'title': video.title,
                'description': video.description,
                'youtube_url': video.youtube_url,
                'embed_url': embed_url,
                'thumbnail_url': video.get_thumbnail_url(),
                'order': video.order,
                'is_short': video.is_youtube_short(),
            })
        
        return Response({
            'success': True,
            'videos': video_list,
            'count': len(video_list)
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in get_active_videos: {str(e)}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== TESTIMONIAL MANAGEMENT ====================

@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def testimonial_list(request):
    """Display list of all testimonials for admin management"""
    try:
        from .models import Testimonial
        
        # Get search query
        search_query = request.GET.get('search', '').strip()
        status_filter = request.GET.get('status', 'all')
        
        # Base queryset
        testimonials = Testimonial.objects.all()
        
        # Apply search filter
        if search_query:
            testimonials = testimonials.filter(
                models.Q(customer_name__icontains=search_query) |
                models.Q(review_text__icontains=search_query) |
                models.Q(location__icontains=search_query)
            )
        
        # Apply status filter
        if status_filter == 'active':
            testimonials = testimonials.filter(is_active=True)
        elif status_filter == 'inactive':
            testimonials = testimonials.filter(is_active=False)
        
        # Order by custom order field
        testimonials = testimonials.order_by('order', '-created_at')
        
        context = {
            'testimonials': testimonials,
            'total_testimonials': Testimonial.objects.count(),
            'active_testimonials': Testimonial.objects.filter(is_active=True).count(),
            'inactive_testimonials': Testimonial.objects.filter(is_active=False).count(),
            'search_query': search_query,
            'status_filter': status_filter,
        }
        
        return render(request, 'testimonials/testimonial_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading testimonials: {str(e)}')
        return redirect('admin_dashboard')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def testimonial_create(request):
    """Create a new testimonial"""
    try:
        from .models import Testimonial
        
        if request.method == 'POST':
            customer_name = request.POST.get('customer_name', '').strip()
            review_text = request.POST.get('review_text', '').strip()
            rating = request.POST.get('rating', 5)
            location = request.POST.get('location', '').strip()
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            customer_image = request.FILES.get('customer_image')
            
            # Validation
            if not customer_name:
                messages.error(request, 'Customer name is required.')
                return redirect('testimonial_create')
            
            if not review_text:
                messages.error(request, 'Review text is required.')
                return redirect('testimonial_create')
            
            # Create testimonial
            testimonial = Testimonial.objects.create(
                customer_name=customer_name,
                review_text=review_text,
                rating=int(rating) if rating else 5,
                location=location if location else None,
                order=int(order) if order else 0,
                is_active=is_active,
                customer_image=customer_image,
                created_by=request.user
            )
            
            messages.success(request, f'Testimonial from "{customer_name}" created successfully!')
            return redirect('testimonial_list')
        
        # GET request - show form
        context = {
            'max_order': Testimonial.objects.count(),
        }
        return render(request, 'testimonials/testimonial_form.html', context)
        
    except Exception as e:
        messages.error(request, f'Error creating testimonial: {str(e)}')
        return redirect('testimonial_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def testimonial_edit(request, testimonial_id):
    """Edit an existing testimonial"""
    try:
        from .models import Testimonial
        
        testimonial = get_object_or_404(Testimonial, id=testimonial_id)
        
        if request.method == 'POST':
            customer_name = request.POST.get('customer_name', '').strip()
            review_text = request.POST.get('review_text', '').strip()
            rating = request.POST.get('rating', 5)
            location = request.POST.get('location', '').strip()
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            customer_image = request.FILES.get('customer_image')
            remove_image = request.POST.get('remove_image') == 'on'
            
            # Validation
            if not customer_name:
                messages.error(request, 'Customer name is required.')
                return redirect('testimonial_edit', testimonial_id=testimonial_id)
            
            if not review_text:
                messages.error(request, 'Review text is required.')
                return redirect('testimonial_edit', testimonial_id=testimonial_id)
            
            # Update testimonial
            testimonial.customer_name = customer_name
            testimonial.review_text = review_text
            testimonial.rating = int(rating) if rating else 5
            testimonial.location = location if location else None
            testimonial.order = int(order) if order else 0
            testimonial.is_active = is_active
            
            # Handle image
            if remove_image and testimonial.customer_image:
                testimonial.customer_image.delete(save=False)
                testimonial.customer_image = None
            elif customer_image:
                # Delete old image if exists
                if testimonial.customer_image:
                    testimonial.customer_image.delete(save=False)
                testimonial.customer_image = customer_image
            
            testimonial.save()
            
            messages.success(request, f'Testimonial from "{customer_name}" updated successfully!')
            return redirect('testimonial_list')
        
        # GET request - show form
        context = {
            'testimonial': testimonial,
            'is_edit': True,
            'max_order': Testimonial.objects.count(),
        }
        return render(request, 'testimonials/testimonial_form.html', context)
        
    except Exception as e:
        messages.error(request, f'Error editing testimonial: {str(e)}')
        return redirect('testimonial_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def testimonial_delete(request, testimonial_id):
    """Delete a testimonial"""
    try:
        from .models import Testimonial
        
        testimonial = get_object_or_404(Testimonial, id=testimonial_id)
        
        if request.method == 'POST':
            name = testimonial.customer_name
            # Delete image file if exists
            if testimonial.customer_image:
                testimonial.customer_image.delete(save=False)
            testimonial.delete()
            messages.success(request, f'Testimonial from "{name}" deleted successfully!')
            return redirect('testimonial_list')
        
        # GET request - show confirmation
        context = {
            'testimonial': testimonial,
        }
        return render(request, 'testimonials/testimonial_confirm_delete.html', context)
        
    except Exception as e:
        messages.error(request, f'Error deleting testimonial: {str(e)}')
        return redirect('testimonial_list')


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
@csrf_exempt
def testimonial_toggle_active(request, testimonial_id):
    """Toggle testimonial active status via AJAX"""
    try:
        from .models import Testimonial
        
        if request.method == 'POST':
            testimonial = get_object_or_404(Testimonial, id=testimonial_id)
            testimonial.is_active = not testimonial.is_active
            testimonial.save()
            
            return JsonResponse({
                'success': True,
                'is_active': testimonial.is_active,
                'message': f'Testimonial is now {"active" if testimonial.is_active else "inactive"}'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
@csrf_exempt
def testimonial_reorder(request):
    """Update testimonial order via AJAX"""
    try:
        from .models import Testimonial
        
        if request.method == 'POST':
            data = json.loads(request.body)
            testimonial_orders = data.get('testimonial_orders', [])
            
            # Update each testimonial's order
            for item in testimonial_orders:
                testimonial_id = item.get('id')
                new_order = item.get('order')
                
                testimonial = Testimonial.objects.get(id=testimonial_id)
                testimonial.order = new_order
                testimonial.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Testimonial order updated successfully'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# Public API for landing page testimonials
@api_view(['GET'])
@permission_classes([])
@authentication_classes([])
def get_active_testimonials(request):
    """Public API endpoint to get active testimonials for landing page carousel"""
    try:
        from .models import Testimonial
        import logging
        logger = logging.getLogger(__name__)
        
        testimonials = Testimonial.objects.filter(is_active=True).order_by('order', '-created_at')
        
        testimonial_list = []
        for testimonial in testimonials:
            testimonial_list.append({
                'id': str(testimonial.id),
                'customer_name': testimonial.customer_name,
                'review_text': testimonial.review_text,
                'rating': testimonial.rating,
                'location': testimonial.location,
                'image_url': testimonial.customer_image.url if testimonial.customer_image else None,
                'order': testimonial.order,
            })
        
        return Response({
            'success': True,
            'testimonials': testimonial_list,
            'count': len(testimonial_list)
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in get_active_testimonials: {str(e)}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@require_POST
def update_patient_status(request, patient_id):
    """Update patient status (green, orange, red)"""
    try:
        data = json.loads(request.body)
        status_value = data.get('status')
        
        if status_value not in ['green', 'orange', 'red']:
            return JsonResponse({'success': False, 'error': 'Invalid status value'}, status=400)
        
        patient = get_object_or_404(Patient, id=patient_id)
        patient.status = status_value
        patient.save()
        
        return JsonResponse({'success': True, 'status': status_value})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def update_patient_sticky_note(request, patient_id):
    """Update patient sticky note"""
    try:
        data = json.loads(request.body)
        sticky_note = data.get('sticky_note', '')
        
        if len(sticky_note) > 500:
            return JsonResponse({'success': False, 'error': 'Sticky note too long (max 500 characters)'}, status=400)
        
        patient = get_object_or_404(Patient, id=patient_id)
        patient.sticky_note = sticky_note
        patient.save()
        
        return JsonResponse({'success': True, 'sticky_note': sticky_note})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def get_patient_sticky_note(request, patient_id):
    """Get patient sticky note"""
    try:
        patient = get_object_or_404(Patient, id=patient_id)
        return JsonResponse({'success': True, 'sticky_note': patient.sticky_note or ''})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
def lab_upload(request):
    """View to handle lab document uploads"""
    if request.method == 'POST':
        try:
            patient_name = request.POST.get('patient_name')
            document = request.FILES.get('document')
            
            if not patient_name or not document:
                messages.error(request, 'Please provide both patient name and a document.')
                return render(request, 'lab_upload.html')
            
            from .models import LabDocument
            LabDocument.objects.create(
                patient_name=patient_name,
                document=document
            )
            
            messages.success(request, 'Lab document uploaded successfully!')
            return redirect('lab_upload')
            
        except Exception as e:
            messages.error(request, f'Error uploading document: {str(e)}')
            
    return render(request, 'lab_upload.html')


# Lab API Views

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_lab_structure(request):
    """Return the structure of lab categories and tests"""
    categories = LabCategory.objects.all().prefetch_related('tests')
    data = []
    for cat in categories:
        tests = []
        for test in cat.tests.all():
            tests.append({
                'id': test.id,
                'name': test.name,
                'unit': test.unit,
                'min': test.min_range,
                'max': test.max_range
            })
        data.append({
            'id': cat.id,
            'name': cat.name,
            'slug': cat.slug,
            'tests': tests
        })
    return JsonResponse({'success': True, 'categories': data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_patient_labs(request, patient_id):
    """Return all lab results for a patient"""
    try:
        patient = Patient.objects.get(id=patient_id)
    except Patient.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Patient not found'}, status=404)
        
    # Check permissions (add your own logic here if needed, e.g. moderator assignment)
    
    results = LabResult.objects.filter(patient=patient).select_related('test', 'test__category').order_by('-date_recorded')
    
    data = []
    for res in results:
        data.append({
            'id': res.id,
            'test_id': res.test.id,
            'test_name': res.test.name,
            'category_name': res.test.category.name,
            'value': res.value,
            'unit': res.test.unit,
            'date_recorded': res.date_recorded.strftime('%Y-%m-%d %H:%M'),
            'recorded_by': res.recorded_by.username if res.recorded_by else 'Unknown',
            'notes': res.notes
        })
        
    return JsonResponse({'success': True, 'results': data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def save_lab_result(request):
    """Save a single lab result"""
    try:
        patient_id = request.data.get('patient_id')
        test_id = request.data.get('test_id')
        value = request.data.get('value')
        date_str = request.data.get('date') # ISO format expected or YYYY-MM-DD
        notes = request.data.get('notes', '')
        
        if not all([patient_id, test_id, value, date_str]):
            return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)
            
        patient = Patient.objects.get(id=patient_id)
        test = LabTest.objects.get(id=test_id)
        
        # Parse date
        # Assuming date_str is suitable for Django DateTimeField or parse it
        from django.utils.dateparse import parse_datetime
        date_recorded = parse_datetime(date_str)
        if not date_recorded:
            # Try simple date if datetime fails
            from django.utils.dateparse import parse_date
            import datetime
            d = parse_date(date_str)
            if d:
                date_recorded = datetime.datetime.combine(d, datetime.time.min)
            else:
                 return JsonResponse({'success': False, 'error': 'Invalid date format'}, status=400)

        LabResult.objects.create(
            patient=patient,
            test=test,
            value=value,
            date_recorded=date_recorded,
            recorded_by=request.user,
            notes=notes
        )
        
        return JsonResponse({'success': True, 'message': 'Lab result saved'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
